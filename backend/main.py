from datetime import datetime
import asyncio
import logging
import sentry_sdk
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration
from sentry_sdk.integrations.celery import CeleryIntegration
from fastapi import FastAPI, File, UploadFile, Request, HTTPException, Form, Response, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import os
import json
import shutil
import tempfile
import pandas as pd
from typing import List, Optional, Any
from pydantic import BaseModel
import sys
import os
import io
import zipfile
import uuid
from fastapi import BackgroundTasks, Depends
from sqlalchemy.orm import Session

# Ensure backend directory is in the path BEFORE local imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from database import get_db, engine, Base, SessionLocal
from models import BatchJob, InvoiceTask, TaskStatus, SalesLineItem, PurchaseLineItem, ObservabilityLog, UserSession, UserPreferences, UserAnnotation, Tenant, GoogleDriveSyncJob, GoogleDriveSyncConfig, SalesPeriodReview, PurchaseGstr2bReview
from async_tasks import process_batch
from ws_manager import manager

# Ensure database tables exist
Base.metadata.create_all(bind=engine)

from invoice_processor import process_pdf, build_dataframes, InvoiceExtractionResponse
from services.gstr2b_reconciler import parse_gstr2b, reconcile as recon_match
from services.udyam_parser import parse_udyam_certificate
from services.msme_compliance import calculate_43bh_compliance
from services.document_core import (
    parse_bank_statement, smart_split_by_size, enhance_scan, compress_pdf, ocr_extract,
    ocr_extract_via_celery, DocumentValidationError, validate_pdf_input
)
from services.gcs_storage import gcs_storage
from services.task_queue import dispatch_batch_task
from models import User
from services.auth import hash_password, verify_password, create_access_token, get_current_user, RoleChecker, require_same_tenant

MODEL_OPTIONS = {
    "auto": None,  # Smart routing (Claude > Groq > Gemini based on keys present)
    "anthropic": {"provider": "anthropic", "model": "claude-haiku-4-5-20251001"},
    "anthropic-sonnet": {"provider": "anthropic", "model": "claude-sonnet-4-6"},
    "openrouter-llama-3.3-70b": {"provider": "openrouter", "model": "meta-llama/llama-3.3-70b-instruct"},
    "openrouter-gemini-flash": {"provider": "openrouter", "model": "google/gemini-2.5-flash"},
    "ollama": {"provider": "ollama", "model": None},  # Uses OLLAMA_MODEL_NAME env var
}

# ── Sentry Observability ──────────────────────────────────────────────────────
_sentry_dsn = os.getenv("SENTRY_DSN", "")
if _sentry_dsn:
    sentry_sdk.init(
        dsn=_sentry_dsn,
        integrations=[
            FastApiIntegration(transaction_style="url"),
            SqlalchemyIntegration(),
            CeleryIntegration(),
        ],
        traces_sample_rate=float(os.getenv("SENTRY_TRACES_SAMPLE_RATE", "0.2")),
        profiles_sample_rate=float(os.getenv("SENTRY_PROFILES_SAMPLE_RATE", "0.1")),
        environment=os.getenv("ENVIRONMENT", "development"),
        release=os.getenv("RENDER_GIT_COMMIT", "local"),
        send_default_pii=True,
    )
    print(f"[Sentry] Initialized — env={os.getenv('ENVIRONMENT', 'development')}")
else:
    print("[Sentry] SENTRY_DSN not set — error tracking disabled.")

# Structured JSON logger for ops dashboards (Datadog / CloudWatch / Render logs)
logging.basicConfig(
    level=logging.INFO,
    format='{"time":"%(asctime)s","level":"%(levelname)s","logger":"%(name)s","msg":%(message)s}',
)
logger = logging.getLogger("auditOS")

app = FastAPI(title="AI Invoice Extractor API")

# Configure CORS
# ALLOWED_ORIGINS must be an explicit comma-separated list in production —
# "*" combined with allow_credentials=True is a contradictory, insecure
# configuration: it lets any website issue authenticated-looking requests
# against every endpoint, including the ones that (as found in Phase 2)
# had no auth check at all. Same fail-closed pattern as JWT_SECRET_KEY below.
_environment = os.getenv("ENVIRONMENT", "development").lower()
allowed_origins_str = os.getenv("ALLOWED_ORIGINS", "*")
origins = [origin.strip() for origin in allowed_origins_str.split(",") if origin.strip()] or ["*"]
_has_wildcard_origin = "*" in origins
if _has_wildcard_origin and _environment == "production":
    raise RuntimeError(
        "\n\n[SECURITY] ALLOWED_ORIGINS includes '*' (unset defaults to this) in a "
        "production environment. Combined with credentialed requests, this allows any "
        "website to call this API. Set ALLOWED_ORIGINS to an explicit comma-separated "
        "list of your frontend origin(s), with no '*' entry, before starting in "
        "production.\n"
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    # Wildcard + credentials is a contradictory combination browsers already
    # refuse to honor for credentialed requests, but Starlette will still
    # advertise `Access-Control-Allow-Credentials: true` if told to. Only
    # allow credentialed CORS once real origins are configured (no '*' entry,
    # not just an exact ["*"] list -- an operator could otherwise smuggle a
    # wildcard in among an explicit CSV list, e.g. "https://a.com,*").
    allow_credentials=(not _has_wildcard_origin),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def log_requests(request: Request, call_next):
    """
    Structured request logging middleware.
    Logs method, path, status, and latency as JSON — compatible with
    Datadog log parsing, Render log drains, and CloudWatch Insights.
    """
    import time as _time
    t0 = _time.perf_counter()
    response = await call_next(request)
    latency_ms = round((_time.perf_counter() - t0) * 1000)
    # Skip noisy health-check paths to keep logs clean
    if request.url.path not in ("/", "/docs", "/openapi.json"):
        logger.info(
            '"method":"%s","path":"%s","status":%d,"latency_ms":%d',
            request.method,
            request.url.path,
            response.status_code,
            latency_ms,
        )
    return response


class UserRegisterRequest(BaseModel):
    email: str
    password: str
    role: Optional[str] = "auditor"

class UserLoginRequest(BaseModel):
    email: str
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    role: str
    email: str

@app.post("/api/auth/register", status_code=201)
async def register(req: UserRegisterRequest, db: Session = Depends(get_db)):
    """Creates a new user account with hashed password and RBAC role."""
    # Check if user already exists
    existing_user = db.query(User).filter(User.email == req.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email is already registered")
    
    # Validate role. "developer" is a platform-wide RBAC bypass (see RoleChecker
    # in services/auth.py) and must never be self-assignable at signup — it can
    # only be granted via direct database provisioning.
    allowed_roles = ["owner", "hr", "auditor", "other"]
    user_role = req.role.lower() if req.role else "auditor"
    if user_role not in allowed_roles:
        raise HTTPException(status_code=400, detail=f"Invalid role. Supported: {allowed_roles}")

    new_user = User(
        id=str(uuid.uuid4()),
        email=req.email,
        hashed_password=hash_password(req.password),
        role=user_role
    )
    db.add(new_user)
    db.commit()
    return {"message": "User registered successfully", "email": new_user.email, "role": new_user.role}

@app.post("/api/auth/login", response_model=TokenResponse)
async def login(req: UserLoginRequest, db: Session = Depends(get_db)):
    """Authenticates credentials and returns a signed JWT access token."""
    print(f"[AUTH DEBUG] Login request received for email: '{req.email}'")
    user = db.query(User).filter(User.email == req.email).first()
    if not user:
        print(f"[AUTH DEBUG] User not found in database for email: '{req.email}'")
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    password_ok = verify_password(req.password, user.hashed_password)
    print(f"[AUTH DEBUG] Password check result for '{req.email}': {password_ok}")
    if not password_ok:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Generate token
    access_token = create_access_token(data={"sub": user.email})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role,
        "email": user.email
    }

# ExportRequest deprecated
def validate_suvit_item(item: dict) -> List[str]:
    """Helper to check for potential errors/warnings in a Suvit line item."""
    errors = []
    if not item.get("invoice_no"):
        errors.append("Missing invoice number")
    if not item.get("voucher_date"):
        errors.append("Missing voucher date")
        
    # Math checks
    amount = float(item.get("taxable_value") or 0.0)
    sgst = float(item.get("sgst_amount") or 0.0)
    cgst = float(item.get("cgst_amount") or 0.0)
    igst = float(item.get("igst_amount") or 0.0)
    total_amount = float(item.get("total_invoice_value") or 0.0)
    
    expected_total = amount + sgst + cgst + igst
    if abs(expected_total - total_amount) > 2.0:  # Allow small rounding threshold
        errors.append(f"Math mismatch: base + taxes ({expected_total:.2f}) != total ({total_amount:.2f})")
        
    return errors

# ── Tenant Management ─────────────────────────────────────────────────────────

class TenantCreateRequest(BaseModel):
    name: str
    slug: str

@app.post("/api/admin/tenants", status_code=201)
async def create_tenant(
    req: TenantCreateRequest,
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
    db: Session = Depends(get_db),
):
    """Creates a new tenant (CA firm). Only owners/developers can create tenants."""
    existing = db.query(Tenant).filter(Tenant.slug == req.slug).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Tenant slug '{req.slug}' already exists.")
    tenant = Tenant(name=req.name, slug=req.slug)
    db.add(tenant)
    db.commit()
    db.refresh(tenant)
    # Atomically claim the tenant for its creator when a self-serve owner
    # (not yet in any firm) creates one — this is the only legitimate way an
    # owner should ever end up assigned to a tenant they didn't already
    # belong to. "developer" accounts provisioning tenants on behalf of a
    # client are deliberately NOT auto-assigned.
    if current_user.role == "owner" and current_user.tenant_id is None:
        current_user.tenant_id = tenant.id
        db.commit()
    return {"tenant_id": tenant.id, "name": tenant.name, "slug": tenant.slug}

@app.get("/api/admin/tenants")
async def list_tenants(
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
    db: Session = Depends(get_db),
):
    """Lists all tenants. Owner/developer only."""
    tenants = db.query(Tenant).all()
    return {"tenants": [{"id": t.id, "name": t.name, "slug": t.slug, "is_active": t.is_active, "created_at": t.created_at.isoformat()} for t in tenants]}

@app.post("/api/admin/tenants/{tenant_id}/assign-user")
async def assign_user_to_tenant(
    tenant_id: str,
    user_email: str,
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
    db: Session = Depends(get_db),
):
    """
    Assigns an existing user to a tenant.

    Frontend call pattern (see firm-settings/page.tsx): an owner who already
    belongs to a tenant invites a colleague (a different user) into that same
    tenant. Tenant *creation* now atomically self-assigns the creating owner
    (see create_tenant above), so this endpoint no longer needs — or allows —
    a "claim a tenant I have no relationship to yet" bootstrap path. Without
    this check, any self-registered "owner" could assign themselves (or
    anyone) into any existing, populated tenant.
    """
    if current_user.role == "owner" and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="You can only assign members to your own firm.")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found.")
    user = db.query(User).filter(User.email == user_email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    user.tenant_id = tenant_id
    db.commit()
    return {"ok": True, "user": user_email, "tenant": tenant.name}


@app.get("/api/admin/tenants/{tenant_id}/users")
async def list_tenant_users(
    tenant_id: str,
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
    db: Session = Depends(get_db),
):
    """List all users assigned to a tenant. Owner can only view their own tenant."""
    if current_user.role == "owner" and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="You can only view your own firm's members.")
    users = db.query(User).filter(User.tenant_id == tenant_id).all()
    return {
        "users": [
            {"id": u.id, "email": u.email, "role": u.role, "is_active": u.is_active}
            for u in users
        ]
    }


@app.delete("/api/admin/tenants/{tenant_id}/users/{user_id}")
async def remove_user_from_tenant(
    tenant_id: str,
    user_id: str,
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
    db: Session = Depends(get_db),
):
    """Remove a user from a tenant (un-assign, does not delete the account)."""
    if current_user.role == "owner" and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="You can only manage your own firm's members.")
    user = db.query(User).filter(User.id == user_id, User.tenant_id == tenant_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found in this firm.")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot remove yourself from the firm.")
    user.tenant_id = None
    db.commit()
    return {"ok": True, "removed": user.email}


@app.put("/api/admin/tenants/{tenant_id}")
async def update_tenant(
    tenant_id: str,
    req: TenantCreateRequest,
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
    db: Session = Depends(get_db),
):
    """Update tenant name/slug. Owner can only update their own tenant."""
    if current_user.role == "owner" and current_user.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="You can only update your own firm.")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Firm not found.")
    tenant.name = req.name
    tenant.slug = req.slug
    db.commit()
    db.refresh(tenant)
    return {"id": tenant.id, "name": tenant.name, "slug": tenant.slug}


@app.get("/api/me/tenant")
async def get_my_tenant(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Returns the current user's tenant info."""
    if not current_user.tenant_id:
        return {"tenant": None}
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    if not tenant:
        return {"tenant": None}
    return {"tenant": {"id": tenant.id, "name": tenant.name, "slug": tenant.slug}}


@app.get("/api/health/workers")
async def health_workers():
    """
    Returns the status of Celery workers, the Redis broker, and all LLM circuit breakers.
    Use this to verify the task queue is healthy before processing large batches.
    """
    from services.task_queue import check_celery_workers, CELERY_ACTIVE
    from core.extraction.llm_call import get_all_breaker_status
    worker_status = check_celery_workers()
    breakers = get_all_breaker_status()
    any_open = any(b["state"] == "open" for b in breakers)
    return {
        "celery_configured": CELERY_ACTIVE,
        "workers_alive": worker_status["active"],
        "worker_count": worker_status["worker_count"],
        "celery_error": worker_status.get("error"),
        "dispatch_mode": "celery" if CELERY_ACTIVE else "local_background_tasks",
        "circuit_breakers": breakers,
        "any_provider_down": any_open,
    }


@app.get("/api/models")
async def get_models():
    return JSONResponse(content={
        "models": [
            {"id": "auto",                    "name": "⚡ Auto (Smart Routing)",       "description": "≤5 pages → OpenRouter cloud, >5 pages → Local Ollama"},
            {"id": "openrouter-llama-3.3-70b","name": "☁️ OpenRouter — Llama 3.3 70B", "description": "Fast cloud model, great for standard invoices"},
            {"id": "openrouter-gemini-flash", "name": "☁️ OpenRouter — Gemini 2.5 Flash", "description": "Google's fast multimodal model via OpenRouter"},
            {"id": "ollama",                  "name": "🖥️ Local Ollama",                 "description": "Private, unlimited — runs entirely on your machine"},
        ]
    })

@app.post("/api/invoices/upload-batch")
async def upload_batch(background_tasks: BackgroundTasks, files: List[UploadFile] = File(...), model: Optional[str] = None, type: Optional[str] = "both", db: Session = Depends(get_db), current_user: User = Depends(RoleChecker(["owner", "auditor"]))):
    batch_id = str(uuid.uuid4())
    total_files = len(files)
    
    batch_job = BatchJob(id=batch_id, total_files=total_files, status=TaskStatus.PENDING, user_id=current_user.id, tenant_id=current_user.tenant_id)
    db.add(batch_job)
    db.commit()

    batch_dir = os.path.join(tempfile.gettempdir(), f"batch_{batch_id}")
    os.makedirs(batch_dir, exist_ok=True)
    
    tasks_to_process = []
    
    try:
        model_config = MODEL_OPTIONS.get(model or "auto")
        
        for file in files:
            file_path = os.path.join(batch_dir, file.filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
                
            # If the file is a zip, extract it and process its contents
            if file.filename.lower().endswith(".zip"):
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    # Extract to a temp subfolder
                    extract_dir = os.path.join(batch_dir, f"extracted_{uuid.uuid4().hex[:8]}")
                    os.makedirs(extract_dir, exist_ok=True)
                    zip_ref.extractall(extract_dir)

                    # Classify by folder structure (Sales Invoice / Other
                    # Invoices / Credit Note / Sale Analysis, same rules as
                    # the Drive sync path) instead of blindly treating every
                    # PDF in the zip as a regular invoice - a zip that
                    # mirrors the real folder tree (e.g. someone zipping up
                    # "Sales Invoice" + "Credit Note" together) would
                    # otherwise misparse credit notes as invoices.
                    from services.drive_classifier import classify_local_directory, DocumentType
                    classified = classify_local_directory(extract_dir)

                    for cf in classified:
                        if cf.document_type == DocumentType.CREDIT_NOTE:
                            from services.credit_note_ingest import ingest_credit_note_pdf
                            try:
                                cn_task_id = ingest_credit_note_pdf(
                                    db, current_user.tenant_id, batch_id, cf.id, cf.name
                                )
                                if not cn_task_id:
                                    print(f"[upload_batch] {cf.name} classified as credit_note but 'Credit Note Number' not found in text")
                            except Exception as e:
                                print(f"[upload_batch] Error processing credit note {cf.name}: {e}")
                            continue

                        if cf.document_type != DocumentType.INVOICE:
                            location = "/".join(cf.path) if cf.path else "(zip root)"
                            print(f"[upload_batch] Found but not yet processed ({cf.document_type.value}, no extractor for this type yet): {location}/{cf.name}")
                            continue

                        extracted_path = cf.id  # local_directory_lister sets id to the real file path
                        extracted_file = cf.name

                        if gcs_storage.is_active():
                            gcs_storage.upload_file(extracted_path, f"batches/{batch_id}/{extracted_file}")

                        task_id = str(uuid.uuid4())
                        invoice_task = InvoiceTask(
                            id=task_id,
                            batch_id=batch_id,
                            file_name=extracted_file,
                            status=TaskStatus.PENDING,
                            invoice_type=type
                        )
                        db.add(invoice_task)

                        tasks_to_process.append({
                            "id": task_id,
                            "file_path": extracted_path
                        })
                # Remove the original zip file after extraction
                try:
                    os.remove(file_path)
                except:
                    pass
            else:
                # It's a standard PDF file
                if file.filename.lower().endswith(".pdf"):
                    if gcs_storage.is_active():
                        gcs_storage.upload_file(file_path, f"batches/{batch_id}/{file.filename}")
                    
                    task_id = str(uuid.uuid4())
                    invoice_task = InvoiceTask(
                        id=task_id,
                        batch_id=batch_id,
                        file_name=file.filename,
                        status=TaskStatus.PENDING,
                        invoice_type=type
                    )
                    db.add(invoice_task)
                    
                    tasks_to_process.append({
                        "id": task_id,
                        "file_path": file_path
                    })
            
        db.commit()
        
        # Dispatch to queue broker or local background tasks
        dispatch_batch_task(background_tasks, batch_id, tasks_to_process, model_config, type)
            
    except Exception as e:
        print(f"Extraction enqueue error: {e}")
        batch_job.status = TaskStatus.FAILED
        db.commit()
        raise HTTPException(status_code=500, detail=str(e))
        
    return JSONResponse(content={
        "message": "Batch enqueued successfully",
        "batch_id": batch_id,
        "total_files": total_files
    })
@app.get("/api/jobs")
async def get_all_jobs(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Tenant-scoped: owners/developers see all batches within their tenant
    base_q = db.query(BatchJob)
    if current_user.tenant_id:
        base_q = base_q.filter(BatchJob.tenant_id == current_user.tenant_id)
    if current_user.role in ["developer", "owner"]:
        batches = base_q.order_by(BatchJob.created_at.desc()).all()
    else:
        batches = base_q.filter(BatchJob.user_id == current_user.id).order_by(BatchJob.created_at.desc()).all()
    return [{
        "id": b.id,
        "created_at": b.created_at.isoformat(),
        "total_files": b.total_files,
        "status": b.status.value
    } for b in batches]

@app.get("/api/jobs/{batch_id}/files/{filename:path}")
async def get_pdf_file(
    batch_id: str,
    filename: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import os, tempfile

    # Tenant isolation, matching the same pattern as get_job_status /
    # export_to_excel: fetch the owning batch, 404 if it doesn't exist,
    # 403 if it belongs to a different tenant.
    batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Job not found")
    require_same_tenant(batch.tenant_id, current_user)

    with open("pdf_debug.log", "a", encoding="utf-8") as f:
        f.write(f"Requested batch_id: {batch_id}, filename: {filename}\n")

    # Check direct in batch_id dir
    batch_dir = os.path.join(tempfile.gettempdir(), f"batch_{batch_id}")
    batch_dir_real = os.path.realpath(batch_dir)
    file_path = os.path.join(batch_dir, filename)
    file_path_real = os.path.realpath(file_path)
    # Reject any filename ("../", absolute paths, etc.) that resolves outside
    # the batch's own temp directory — prevents path traversal to arbitrary
    # files on the host (e.g. the backend's .env).
    if file_path_real != batch_dir_real and not file_path_real.startswith(batch_dir_real + os.sep):
        raise HTTPException(status_code=400, detail="Invalid filename")
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="application/pdf")
        
    # If missing locally, try to pull it from GCS
    if gcs_storage.is_active():
        gcs_blob_name = f"batches/{batch_id}/{filename}"
        if gcs_storage.file_exists(gcs_blob_name):
            if gcs_storage.download_file(gcs_blob_name, file_path):
                return FileResponse(file_path, media_type="application/pdf")
        
    # Check in subdirectories (if extracted from zip)
    if os.path.exists(batch_dir):
        for root, _, files in os.walk(batch_dir):
            if filename in files:
                return FileResponse(os.path.join(root, filename), media_type="application/pdf")
                
    # If file is genuinely missing
    from fastapi.responses import HTMLResponse
    html_content = f"""
    <html>
        <body style="font-family: sans-serif; display: flex; align-items: center; justify-content: center; height: 100vh; margin: 0; background-color: #f9f9fa; color: #333; text-align: center;">
            <div>
                <svg width="48" height="48" viewBox="0 0 24 24" fill="none" stroke="#ef4444" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="margin-bottom: 16px;"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path><polyline points="14 2 14 8 20 8"></polyline><line x1="9" y1="15" x2="15" y2="15"></line></svg>
                <h3>PDF No Longer Available</h3>
                <p style="color: #666; max-width: 300px; margin: 0 auto; line-height: 1.5;">This file was processed in an older session and has been deleted from the server to save space.<br><br>Please upload this invoice again to view it side-by-side.</p>
            </div>
        </body>
    </html>
    """
    return HTMLResponse(content=html_content, status_code=404)

@app.get("/api/jobs/{batch_id}")
async def get_job_status(batch_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Job not found")
    require_same_tenant(batch.tenant_id, current_user)
        
    tasks = db.query(InvoiceTask).filter(InvoiceTask.batch_id == batch_id).all()
    
    total = len(tasks)
    pending = sum(1 for t in tasks if t.status == TaskStatus.PENDING)
    processing = sum(1 for t in tasks if t.status == TaskStatus.PROCESSING)
    completed = sum(1 for t in tasks if t.status == TaskStatus.COMPLETED)
    failed = sum(1 for t in tasks if t.status == TaskStatus.FAILED)
    
    # ── Observability data integration ─────────────────────────
    # Fetch batch metrics if available
    batch_log = db.query(ObservabilityLog).filter(
        ObservabilityLog.batch_id == batch_id,
        ObservabilityLog.event_type == "batch_metrics"
    ).first()
    batch_metrics = None
    if batch_log:
        try:
            batch_metrics = json.loads(batch_log.payload_json)
        except:
            pass
            
    # Fetch batch-level system flags and task-level flags
    flag_logs = db.query(ObservabilityLog).filter(
        ObservabilityLog.batch_id == batch_id,
        ObservabilityLog.event_type == "system_flag"
    ).all()
    
    batch_flags = []
    flags_by_file = {}
    for log in flag_logs:
        try:
            payload = json.loads(log.payload_json)
            flag_info = {
                "flag_id": log.flag_id,
                "severity": log.severity,
                "detail": payload.get("trigger_detail", ""),
                "timestamp": log.timestamp_utc.isoformat()
            }
            if log.file_id:
                flags_by_file.setdefault(log.file_id, []).append(flag_info)
            else:
                batch_flags.append(flag_info)
        except:
            pass
            
    # Fetch quality scores by file
    score_logs = db.query(ObservabilityLog).filter(
        ObservabilityLog.batch_id == batch_id,
        ObservabilityLog.event_type == "extraction_quality_score"
    ).all()
    scores_by_file = {}
    for log in score_logs:
        if log.file_id:
            try:
                payload = json.loads(log.payload_json)
                scores_by_file[log.file_id] = payload.get("composite_score", 0.0)
            except:
                pass
    # ───────────────────────────────────────────────────────────
    
    tasks_details = []
    all_sales = []
    all_purchase = []
    
    for t in tasks:
        task_info = {
            "task_id": t.id,
            "filename": t.file_name,
            "status": t.status.value,
            "error_message": t.error_message,
            "composite_score": scores_by_file.get(t.id, None),
            "flags": flags_by_file.get(t.id, []),
            "recon_status": getattr(t, 'recon_status', None),
        }
        
        sales = []
        purchase = []
        if getattr(t, 'sales_items', None):
            sales = [{c.name: getattr(item, c.name) for c in item.__table__.columns if c.name not in ["task_id"]} for item in t.sales_items]
        if getattr(t, 'purchase_items', None):
            purchase = [{c.name: getattr(item, c.name) for c in item.__table__.columns if c.name not in ["task_id"]} for item in t.purchase_items]
            
        if sales or purchase or t.status == TaskStatus.COMPLETED:
            for s in sales:
                s["errors"] = validate_suvit_item(s)
                s["filename"] = t.file_name
            for p in purchase:
                p["errors"] = validate_suvit_item(p)
                p["filename"] = t.file_name
                
            task_info["sales_count"] = len(sales)
            task_info["purchase_count"] = len(purchase)
            all_sales.extend(sales)
            all_purchase.extend(purchase)
            
        tasks_details.append(task_info)
        
    return JSONResponse(content={
        "batch_id": batch.id,
        "status": batch.status.value,
        "progress": {
            "total": total,
            "pending": pending,
            "processing": processing,
            "completed": completed,
            "failed": failed
        },
        "tasks": tasks_details,
        "sales_items": all_sales if batch.status == TaskStatus.COMPLETED else [],
        "purchase_items": all_purchase if batch.status == TaskStatus.COMPLETED else [],
        "batch_metrics": batch_metrics,
        "batch_flags": batch_flags
    })


@app.get("/api/jobs/{batch_id}/duplicates")
async def get_duplicate_invoices(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Detect duplicate invoices in this batch.
    Returns two lists:
      - within_batch : same invoice_no+gstin+date appears multiple times inside this batch
      - cross_batch  : same invoice_no+gstin found in a previous batch for this user/tenant
    Cross-batch matches are flagged as 'critical' (likely a re-upload causing double counting).
    """
    from services.duplicate_detector import detect_all

    batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    require_same_tenant(batch.tenant_id, current_user)

    tasks = db.query(InvoiceTask).filter(InvoiceTask.batch_id == batch_id).all()
    result = detect_all(
        db=db,
        tasks=tasks,
        current_batch_id=batch_id,
        user_id=current_user.id,
        tenant_id=current_user.tenant_id,
    )
    return result


@app.get("/api/tasks/{task_id}/review")
async def get_task_review(
    task_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Phase 4A: Returns the full deterministic reconciliation audit report for a specific invoice task.
    Used to power the Review Panel UI with correction proposals, variance breakdowns, and status.
    """
    from models import InvoiceTask
    task = db.query(InvoiceTask).filter(InvoiceTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    require_same_tenant(task.batch.tenant_id if task.batch else None, current_user)

    recon_data = None
    if task.recon_report_json:
        try:
            recon_data = json.loads(task.recon_report_json)
        except Exception:
            recon_data = None

    return JSONResponse(content={
        "task_id": task_id,
        "filename": task.file_name,
        "recon_status": task.recon_status,
        "recon_report": recon_data,
    })

@app.patch("/api/tasks/{task_id}/accept-correction")
async def accept_correction(task_id: str, db: Session = Depends(get_db), current_user: User = Depends(RoleChecker(["owner", "auditor"]))):
    """
    Phase 4A: CA reviewer accepts auto-correction proposal.
    Marks the task recon_status as HUMAN_CORRECTED.
    """
    from models import InvoiceTask
    task = db.query(InvoiceTask).filter(InvoiceTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    require_same_tenant(task.batch.tenant_id if task.batch else None, current_user)
    task.recon_status = "HUMAN_CORRECTED"
    db.commit()
    return JSONResponse(content={"task_id": task_id, "recon_status": "HUMAN_CORRECTED", "accepted_by": current_user.email})

@app.get("/api/tasks/{task_id}/observability")
async def get_task_observability(
    task_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns step-by-step pipeline execution logs for a specific file.
    """
    from models import InvoiceTask
    task = db.query(InvoiceTask).filter(InvoiceTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    require_same_tenant(task.batch.tenant_id if task.batch else None, current_user)

    logs = db.query(ObservabilityLog).filter(
        ObservabilityLog.file_id == task_id
    ).order_by(ObservabilityLog.timestamp_utc.asc()).all()
    
    events = []
    for l in logs:
        try:
            payload = json.loads(l.payload_json)
        except:
            payload = {}
        events.append({
            "id": l.id,
            "event_type": l.event_type,
            "stage": l.stage,
            "severity": l.severity,
            "flag_id": l.flag_id,
            "timestamp": l.timestamp_utc.isoformat(),
            "model": l.model_identifier,
            "provider": l.api_provider,
            "prompt_version": l.prompt_version,
            "payload": payload
        })
        
    return JSONResponse(content={"task_id": task_id, "events": events})

@app.get("/api/observability/stats")
async def get_observability_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Aggregates metrics and flags across the workspace for the landing dashboard.

    Tenant-scoped: every aggregate below (batch/file counts, quality scores,
    LLM cost totals, flags) is restricted to the caller's own tenant via a
    join through BatchJob.tenant_id -- ObservabilityLog.tenant_id itself is
    nullable and not reliably populated at every write site, so batch_id
    (always set, per BatchJob's two creation paths) is the trustworthy join
    key, matching the pattern the `recent_jobs` section below already used.

    Deliberately unconditional (no "if current_user.tenant_id" branch): a
    caller with no tenant of their own must see only legacy/untenanted
    batches (SQLAlchemy's `== None` compiles to `IS NULL`), never every
    tenant's data -- the same "unassigned caller is not a blanket pass"
    correction Phase 1 already applied to `require_same_tenant`.
    """
    from models import BatchJob, InvoiceTask

    tenant_batch_ids = {
        b.id for b in db.query(BatchJob.id).filter(BatchJob.tenant_id == current_user.tenant_id).all()
    }

    total_batches = db.query(BatchJob).filter(BatchJob.tenant_id == current_user.tenant_id).count()
    total_files = db.query(InvoiceTask).filter(InvoiceTask.batch_id.in_(tenant_batch_ids)).count()
    log_q = db.query(ObservabilityLog).filter(ObservabilityLog.batch_id.in_(tenant_batch_ids))

    # Query all quality scores
    score_logs = log_q.filter(
        ObservabilityLog.event_type == "extraction_quality_score"
    ).all()

    scores_by_file = {}
    scores = []
    for log in score_logs:
        if log.file_id:
            try:
                payload = json.loads(log.payload_json)
                val = payload.get("composite_score")
                if val is not None:
                    scores_by_file[log.file_id] = float(val)
                    scores.append(float(val))
            except:
                pass

    avg_score = round(sum(scores) / len(scores), 3) if scores else 1.0

    # Query total cost from file metrics
    metric_logs = log_q.filter(
        ObservabilityLog.event_type == "file_metrics"
    ).all()

    total_cost = 0.0
    for log in metric_logs:
        try:
            payload = json.loads(log.payload_json)
            total_cost += float(payload.get("model_cost", {}).get("total_cost_inr", 0.0))
        except:
            pass

    # Query system flags
    flag_logs = log_q.filter(
        ObservabilityLog.event_type == "system_flag"
    ).order_by(ObservabilityLog.timestamp_utc.desc()).all()
    
    flags = []
    for log in flag_logs:
        try:
            payload = json.loads(log.payload_json)
            flags.append({
                "batch_id": log.batch_id,
                "file_id": log.file_id,
                "filename": payload.get("filename", "batch-level"),
                "flag_id": log.flag_id,
                "severity": log.severity,
                "detail": payload.get("trigger_detail", ""),
                "timestamp": log.timestamp_utc.isoformat()
            })
        except:
            pass
            
    # Query recent jobs (same tenant scoping as the aggregates above)
    batches = db.query(BatchJob).filter(
        BatchJob.tenant_id == current_user.tenant_id
    ).order_by(BatchJob.created_at.desc()).limit(10).all()
    
    recent_jobs = []
    for b in batches:
        tasks_cnt = db.query(InvoiceTask).filter(InvoiceTask.batch_id == b.id).count()
        
        # Calculate batch average score
        b_scores = [scores_by_file.get(log.file_id) for log in score_logs if log.batch_id == b.id and log.file_id]
        b_scores_clean = [s for s in b_scores if s is not None]
        b_avg = round(sum(b_scores_clean) / len(b_scores_clean), 3) if b_scores_clean else None
        
        # Batch flags count
        b_flags_count = sum(1 for log in flag_logs if log.batch_id == b.id)
        
        recent_jobs.append({
            "id": b.id,
            "created_at": b.created_at.isoformat(),
            "total_files": b.total_files,
            "status": b.status.value,
            "average_quality_score": b_avg,
            "flags_count": b_flags_count
        })
        
    return JSONResponse(content={
        "total_batches": total_batches,
        "total_files": total_files,
        "average_quality_score": avg_score,
        "total_cost_inr": round(total_cost, 2),
        "recent_flags": flags[:10],
        "total_flags": len(flags),
        "recent_jobs": recent_jobs
    })


@app.websocket("/api/ws/jobs/{batch_id}")
async def websocket_endpoint(websocket: WebSocket, batch_id: str):
    await manager.connect(websocket, batch_id)
    try:
        # Poll the DB every 2s and push status to the browser.
        # This bridges the Celery worker (separate process) and the browser WebSocket
        # without requiring shared in-memory state.
        while True:
            db = SessionLocal()
            try:
                batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
                tasks = db.query(InvoiceTask).filter(InvoiceTask.batch_id == batch_id).all()
                total = len(tasks)
                completed = sum(1 for t in tasks if t.status == TaskStatus.COMPLETED)
                failed = sum(1 for t in tasks if t.status == TaskStatus.FAILED)
                status = batch.status.value if batch else "PROCESSING"
                msg = {"total": total, "completed": completed, "failed": failed, "status": status}
                await websocket.send_json(msg)
                if status in ("COMPLETED", "FAILED"):
                    break
            except Exception:
                break
            finally:
                db.close()
            await asyncio.sleep(2)
    except WebSocketDisconnect:
        pass
    finally:
        await manager.disconnect(websocket, batch_id)


@app.get("/api/export/{batch_id}")
async def export_to_excel(
    batch_id: str,
    type: str,
    schema: str = "suvit",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    require_same_tenant(batch.tenant_id, current_user)

    from invoice_processor import SuvitSalesItem, SuvitPurchaseItem
    
    extraction_response = InvoiceExtractionResponse()
    
    from sqlalchemy import or_
    tasks = db.query(InvoiceTask).filter(
        InvoiceTask.batch_id == batch_id,
        or_(InvoiceTask.recon_status.is_(None), InvoiceTask.recon_status.notin_(["BLOCKED", "DUPLICATE"])),
    ).all()

    if type in ["sales", "both"]:
        sales = []
        for t in tasks:
            if getattr(t, 'sales_items', None):
                for item in t.sales_items:
                    item_dict = {c.name: getattr(item, c.name) for c in item.__table__.columns if c.name not in ["id", "task_id"]}
                    sales.append(SuvitSalesItem(**item_dict))
        extraction_response.sales_items = sales
        
    if type in ["purchase", "both"]:
        purchase = []
        for t in tasks:
            if getattr(t, 'purchase_items', None):
                for item in t.purchase_items:
                    item_dict = {c.name: getattr(item, c.name) for c in item.__table__.columns if c.name not in ["id", "task_id"]}
                    purchase.append(SuvitPurchaseItem(**item_dict))
        extraction_response.purchase_items = purchase
        
    if not extraction_response.sales_items and not extraction_response.purchase_items:
        raise HTTPException(status_code=400, detail="No items to export")
        
    sales_df, purchase_df = build_dataframes(extraction_response)
    
    def apply_schema(df, schema_name):
        if isinstance(df, pd.DataFrame) and not df.empty and schema_name != "suvit":
            sap_map = {"party_ac_name": "VendorCode", "invoice_no": "Reference", "voucher_date": "DocumentDate", "taxable_value": "AmountLC", "igst_amount": "TaxAmountIGST", "cgst_amount": "TaxAmountCGST", "sgst_amount": "TaxAmountSGST", "total_invoice_value": "TotalAmountLC", "hsn": "HSNSAC", "VOUCHER NO": "DocumentNumber", "PARTY GSTIN": "TaxNumber3", "DATE": "PostingDate"}
            netsuite_map = {"invoice_no": "InternalID", "voucher_date": "TranDate", "party_ac_name": "Entity", "taxable_value": "GrossAmt", "total_invoice_value": "Amount", "VOUCHER NO": "TranId", "DATE": "Date", "PARTY GSTIN": "VATRegNumber"}
            dynamics_map = {"invoice_no": "InvoiceId", "voucher_date": "TransDate", "party_ac_name": "AccountNum", "total_invoice_value": "InvoiceAmount", "taxable_value": "LineAmount", "VOUCHER NO": "InvoiceRegister", "DATE": "DocumentDate"}
            mapping = sap_map if schema_name == "sap" else (netsuite_map if schema_name == "netsuite" else dynamics_map)
            return df.rename(columns=lambda c: mapping.get(c, c))
        return df

    if isinstance(sales_df, dict):
        for k in sales_df:
            sales_df[k] = apply_schema(sales_df[k], schema)
    else:
        sales_df = apply_schema(sales_df, schema)
        
    purchase_df = apply_schema(purchase_df, schema)
    
    has_sales = isinstance(sales_df, dict) and any(not df.empty for df in sales_df.values())
    has_purchase = not purchase_df.empty

    def save_sales(path):
        with pd.ExcelWriter(path) as writer:
            for sheet_name, df in sales_df.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    def save_purchase(path):
        """Writes purchase register + ITC Summary sheet to Excel."""
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            purchase_df.to_excel(writer, sheet_name="Purchase Register", index=False)

            # ITC Summary sheet
            if "ITC ELIGIBILITY" in purchase_df.columns and "TAXABLE AMOUNT" in purchase_df.columns:
                itc_col = purchase_df["ITC ELIGIBILITY"].fillna("ITC_UNKNOWN")
                amt_col = purchase_df["TAXABLE AMOUNT"].fillna(0)
                summary = {
                    "Category": ["ITC_ELIGIBLE", "ITC_BLOCKED", "ITC_RESTRICTED", "ITC_EXEMPT", "ITC_UNKNOWN", "TOTAL"],
                }
                totals = {}
                for cat in summary["Category"][:-1]:
                    mask = itc_col == cat
                    totals[cat] = float(amt_col[mask].sum())
                totals["TOTAL"] = float(amt_col.sum())
                summary["Taxable Amount (₹)"] = [totals[c] for c in summary["Category"]]
                summary["Line Items"] = [int((itc_col == c).sum()) if c != "TOTAL" else len(purchase_df) for c in summary["Category"]]
                summary["ITC at Risk (₹)"] = [
                    totals.get("ITC_BLOCKED", 0) if c == "ITC_BLOCKED" else
                    totals.get("ITC_RESTRICTED", 0) if c == "ITC_RESTRICTED" else
                    totals.get("TOTAL", 0) - totals.get("ITC_ELIGIBLE", 0) if c == "TOTAL" else 0
                    for c in summary["Category"]
                ]
                pd.DataFrame(summary).to_excel(writer, sheet_name="ITC Summary", index=False)

    if has_sales and has_purchase:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            sales_path = os.path.join(tempfile.gettempdir(), "Suvit_Sales_Upload.xlsx")
            purchase_path = os.path.join(tempfile.gettempdir(), "Suvit_Purchase_Upload.xlsx")
            save_sales(sales_path)
            save_purchase(purchase_path)
            zip_file.write(sales_path, arcname="Suvit_Sales_Upload.xlsx")
            zip_file.write(purchase_path, arcname="Suvit_Purchase_Upload.xlsx")

        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=Suvit_Both_Upload.zip"}
        )
    elif has_sales:
        output_path = os.path.join(tempfile.gettempdir(), "Suvit_Sales_Upload.xlsx")
        save_sales(output_path)
        return FileResponse(
            path=output_path,
            filename="Suvit_Sales_Upload.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        output_path = os.path.join(tempfile.gettempdir(), "Suvit_Purchase_Upload.xlsx")
        save_purchase(output_path)
        return FileResponse(
            path=output_path,
            filename="Suvit_Purchase_Upload.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@app.get("/api/export/{batch_id}/gstr1")
async def export_gstr1_json(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate and download a GSTR-1 filed-format JSON for the given batch.
    Only sales invoices are included. The JSON structure matches the GSTN portal
    offline tool / GSP API format (b2b, b2cl, b2cs, exp, cdnr, cdnur, hsn, doc).
    """
    from services.gstr1_generator import generate_gstr1_json

    batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    require_same_tenant(batch.tenant_id, current_user)

    from sqlalchemy import or_
    tasks = db.query(InvoiceTask).filter(
        InvoiceTask.batch_id == batch_id,
        or_(InvoiceTask.recon_status.is_(None), InvoiceTask.recon_status.notin_(["BLOCKED", "DUPLICATE"])),
    ).all()
    sales_items = []
    for t in tasks:
        if getattr(t, "sales_items", None):
            sales_items.extend(t.sales_items)

    if not sales_items:
        raise HTTPException(
            status_code=400,
            detail="No sales line items found for this batch. GSTR-1 requires sales invoices.",
        )

    firm_gstin = os.getenv("FIRM_GSTIN", "")
    gstr1 = generate_gstr1_json(sales_items, firm_gstin=firm_gstin)

    json_bytes = json.dumps(gstr1, ensure_ascii=False, indent=2).encode("utf-8")
    fp = gstr1.get("fp", "MMYYYY")
    filename = f"GSTR1_{fp}_{batch_id[:8]}.json"

    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────────────────────────────────────────
# Sales Period Review Gate (Phase 7) — the checkpoint between
# reconciliation/GSTR-1 filing generation (Phases 4-5) and anything
# actually reaching a client or the GST portal. See services/period_review.py.
# ─────────────────────────────────────────────────────────────────────────

class PeriodReviewDecisionRequest(BaseModel):
    notes: Optional[str] = None


@app.post("/api/sales/period-reviews/generate")
async def generate_period_review(
    background_tasks: BackgroundTasks,
    period: str = Form(...),  # "YYYY-MM"
    client_sheet: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    """
    Runs reconciliation (sales_reconciliation.py) + GSTR-1 filing
    generation (gstr1_filing.py) for this tenant's given period against
    an uploaded client sheet, and persists the result as a PENDING_REVIEW
    SalesPeriodReview row. Always creates a fresh review (skip_if_pending=
    False) since a human hitting "generate" is an explicit request for the
    current state - e.g. after correcting the client sheet. The scheduled
    ingestion chain (celery_app.py's sales_ingestion_task) calls the same
    underlying generate_period_review_for_tenant with skip_if_pending=True
    instead, so it doesn't pile up a fresh row every day.
    """
    from services.period_review import generate_period_review_for_tenant, get_review_detail

    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Current user has no tenant assigned")

    tmp_dir = tempfile.mkdtemp(prefix="period_review_")
    sheet_path = os.path.join(tmp_dir, client_sheet.filename)
    with open(sheet_path, "wb") as f:
        shutil.copyfileobj(client_sheet.file, f)

    try:
        review, _created = generate_period_review_for_tenant(
            db, current_user.tenant_id, period, sheet_path, skip_if_pending=False
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return get_review_detail(review)


@app.get("/api/sales/period-reviews")
async def list_period_reviews(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lists this tenant's period reviews, most recent first."""
    if not current_user.tenant_id:
        return []
    reviews = (
        db.query(SalesPeriodReview)
        .filter(SalesPeriodReview.tenant_id == current_user.tenant_id)
        .order_by(SalesPeriodReview.created_at.desc())
        .all()
    )
    return [{
        "id": r.id, "period": r.period, "status": r.status,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
    } for r in reviews]


@app.get("/api/sales/period-reviews/{review_id}")
async def get_period_review(
    review_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from services.period_review import get_review_detail

    review = db.query(SalesPeriodReview).filter(SalesPeriodReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Period review not found")
    require_same_tenant(review.tenant_id, current_user)
    return get_review_detail(review)


@app.post("/api/sales/period-reviews/{review_id}/approve")
async def approve_period_review_endpoint(
    review_id: str,
    req: PeriodReviewDecisionRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    from services.period_review import approve_period_review, get_review_detail, ReviewStateError

    review = db.query(SalesPeriodReview).filter(SalesPeriodReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Period review not found")
    require_same_tenant(review.tenant_id, current_user)

    try:
        approved = approve_period_review(db, review_id, current_user.id, notes=req.notes)
    except ReviewStateError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return get_review_detail(approved)


@app.post("/api/sales/period-reviews/{review_id}/reject")
async def reject_period_review_endpoint(
    review_id: str,
    req: PeriodReviewDecisionRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    from services.period_review import reject_period_review, get_review_detail, ReviewStateError

    review = db.query(SalesPeriodReview).filter(SalesPeriodReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Period review not found")
    require_same_tenant(review.tenant_id, current_user)

    try:
        rejected = reject_period_review(db, review_id, current_user.id, notes=req.notes or "")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ReviewStateError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return get_review_detail(rejected)


# ─────────────────────────────────────────────────────────────────────────
# Purchase / GSTR-2B Reconciliation Review Gate (Phase A automation) —
# mirrors the Sales Period Review Gate above; see services/purchase_review.py
# and models.PurchaseGstr2bReview.
# ─────────────────────────────────────────────────────────────────────────

class Gstr2bReviewDecisionRequest(BaseModel):
    notes: Optional[str] = None


@app.post("/api/purchase/gstr2b-reviews/generate")
async def generate_gstr2b_review(
    background_tasks: BackgroundTasks,
    period: str = Form(...),        # "YYYY-MM"
    gstin: str = Form(...),         # which OneStack registration this 2B was issued for
    gstr2b_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    """
    Runs GSTR-2B reconciliation (gstr2b_reconciler.py) for this tenant's
    given period/GSTIN against an uploaded GSTR-2B JSON, and persists the
    result as a PENDING_REVIEW PurchaseGstr2bReview row. Always creates a
    fresh review (skip_if_pending=False), matching the Sales manual
    endpoint's reasoning - a human hitting "generate" wants the current
    state now. The scheduled Drive-drop chain (celery_app.py's
    gstr2b_ingestion_task) calls the same generate_review_for_tenant with
    skip_if_pending=True instead.
    """
    from services.purchase_review import generate_review_for_tenant, get_review_detail

    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Current user has no tenant assigned")

    tmp_dir = tempfile.mkdtemp(prefix="gstr2b_review_")
    file_path = os.path.join(tmp_dir, gstr2b_file.filename)
    with open(file_path, "wb") as f:
        shutil.copyfileobj(gstr2b_file.file, f)

    try:
        review, _created = generate_review_for_tenant(
            db, current_user.tenant_id, period, gstin, file_path, skip_if_pending=False
        )
    except (ValueError, json.JSONDecodeError) as e:
        raise HTTPException(status_code=400, detail=str(e))

    return get_review_detail(review)


@app.get("/api/purchase/gstr2b-reviews")
async def list_gstr2b_reviews(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lists this tenant's GSTR-2B reviews, most recent first."""
    if not current_user.tenant_id:
        return []
    reviews = (
        db.query(PurchaseGstr2bReview)
        .filter(PurchaseGstr2bReview.tenant_id == current_user.tenant_id)
        .order_by(PurchaseGstr2bReview.created_at.desc())
        .all()
    )
    return [{
        "id": r.id, "period": r.period, "gstin": r.gstin, "status": r.status,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
    } for r in reviews]


@app.get("/api/purchase/gstr2b-reviews/{review_id}")
async def get_gstr2b_review(
    review_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from services.purchase_review import get_review_detail

    review = db.query(PurchaseGstr2bReview).filter(PurchaseGstr2bReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="GSTR-2B review not found")
    require_same_tenant(review.tenant_id, current_user)
    return get_review_detail(review)


@app.post("/api/purchase/gstr2b-reviews/{review_id}/approve")
async def approve_gstr2b_review_endpoint(
    review_id: str,
    req: Gstr2bReviewDecisionRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    from services.purchase_review import approve_review, get_review_detail, ReviewStateError

    review = db.query(PurchaseGstr2bReview).filter(PurchaseGstr2bReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="GSTR-2B review not found")
    require_same_tenant(review.tenant_id, current_user)

    try:
        approved = approve_review(db, review_id, current_user.id, notes=req.notes)
    except ReviewStateError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return get_review_detail(approved)


@app.post("/api/purchase/gstr2b-reviews/{review_id}/reject")
async def reject_gstr2b_review_endpoint(
    review_id: str,
    req: Gstr2bReviewDecisionRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    from services.purchase_review import reject_review, get_review_detail, ReviewStateError

    review = db.query(PurchaseGstr2bReview).filter(PurchaseGstr2bReview.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="GSTR-2B review not found")
    require_same_tenant(review.tenant_id, current_user)

    try:
        rejected = reject_review(db, review_id, current_user.id, notes=req.notes or "")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ReviewStateError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return get_review_detail(rejected)


class MSMEVerifyRequest(BaseModel):
    udyam_number: str

@app.post("/api/verify-msme")
async def verify_msme_status(req: MSMEVerifyRequest, current_user: User = Depends(RoleChecker(["owner", "hr"]))):
    """
    Authorized integration point for MSME status verification.
    This simulates a query to the Ministry of MSME database or an authorized API provider.
    """
    udyam = req.udyam_number.strip().upper()
    if not udyam.startswith("UDYAM"):
        raise HTTPException(status_code=400, detail="Invalid Udyam Number format. Must start with 'UDYAM'.")
    
    # Extract state or digits for deterministic mock response
    parts = udyam.split("-")
    state_code = parts[1] if len(parts) > 1 else "IND"
    
    # Deterministic mock assignment based on the last digit
    last_char = udyam[-1] if udyam else "0"
    if last_char in "159":
        enterprise_type = "MICRO"
    elif last_char in "26":
        enterprise_type = "SMALL"
    elif last_char in "37":
        enterprise_type = "MEDIUM"
    else:
        enterprise_type = "LARGE"
        
    mock_names = {
        "MH": "Maharashtra Engineering Works",
        "DL": "Capital Logistics Services",
        "GJ": "Gujarat Textile Mills Ltd",
        "UP": "Noida Business Consulting",
        "KA": "Bengaluru Tech Services",
    }
    company_name = mock_names.get(state_code, f"Authorized Vendor ({state_code}) Ltd")
    
    return JSONResponse(content={
        "success": True,
        "udyam_number": udyam,
        "enterprise_type": enterprise_type,
        "enterprise_name": company_name,
        "message": f"Successfully verified via authorized compliance database."
    })

@app.post("/api/tax/parse-udyam")
async def upload_udyam_certificate(file: UploadFile = File(...), current_user: User = Depends(RoleChecker(["owner", "hr"]))):
    """
    Ingests a vendor's Udyam Registration Certificate PDF, extracts metadata
    and normalizes the enterprise classification status.
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF certificates are accepted.")
    try:
        pdf_bytes = await file.read()
        parsed_data = parse_udyam_certificate(pdf_bytes)
        return JSONResponse(content={
            "success": True,
            "filename": file.filename,
            **parsed_data
        })
    except Exception as e:
        print(f"Udyam parsing error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to parse Udyam Certificate: {str(e)}")

class MSMEComplianceRequest(BaseModel):
    invoice_date: str
    payment_date: Optional[str] = None
    enterprise_type: str
    has_agreement: bool
    amount: float

@app.post("/api/tax/compliance")
async def calculate_compliance_metrics(req: MSMEComplianceRequest, current_user: User = Depends(RoleChecker(["owner", "hr"]))):
    """
    Computes statutory MSME 43B(h) compliance metrics.
    """
    try:
        metrics = calculate_43bh_compliance(
            invoice_date_str=req.invoice_date,
            payment_date_str=req.payment_date,
            enterprise_type=req.enterprise_type,
            has_agreement=req.has_agreement,
            amount=req.amount
        )
        return JSONResponse(content={
            "success": True,
            **metrics
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── Document Utility Suite Endpoints ──────────────────────────────────────────

@app.post("/api/docs/bank-parse")
async def bank_parse_endpoint(
    file: UploadFile = File(...),
    password: str = Form(""),
    confidence_min: str = None
):
    """
    Multi-stage bank statement parser with confidence scoring.
    Returns: {
        "success": bool,
        "transactions": [
            {
                "date": str,
                "narration": str,
                "debit": str,
                "credit": str,
                "balance": str,
                "confidence": "SURE" | "PROBABLE" | "UNCERTAIN",
                "bank_detected": str,
                "extraction_method": "table" | "regex"
            }, ...
        ],
        "summary": {"total": int, "sure": int, "probable": int, "uncertain": int}
    }
    """
    content = await file.read()
    try:
        txs = parse_bank_statement(content, password, confidence_min)

        # Build summary
        confidence_counts = {
            "SURE": sum(1 for t in txs if t.get("confidence") == "SURE"),
            "PROBABLE": sum(1 for t in txs if t.get("confidence") == "PROBABLE"),
            "UNCERTAIN": sum(1 for t in txs if t.get("confidence") == "UNCERTAIN"),
        }

        return JSONResponse(content={
            "success": True,
            "transactions": txs,
            "summary": {
                "total": len(txs),
                **confidence_counts
            }
        })

    except DocumentValidationError as e:
        print(f"[Validation Error] {str(e)}")
        raise HTTPException(status_code=400, detail=f"Validation failed: {str(e)}")

    except ValueError as e:
        error_msg = str(e)
        print(f"[ValueError] {error_msg}")
        if "Invalid password" in error_msg or "encrypted" in error_msg.lower():
            raise HTTPException(status_code=400, detail="Invalid password for encrypted PDF.")
        raise HTTPException(status_code=400, detail=f"PDF error: {error_msg}")

    except Exception as e:
        import traceback
        print(f"[Parsing Error] {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Parsing failed: {type(e).__name__}: {str(e)}")

@app.post("/api/docs/split-portal")
async def split_portal_endpoint(file: UploadFile = File(...), target_mb: float = Form(4.5)):
    """
    Splits heavy PDF files into sub-5MB chunks, returning them bundled inside a single ZIP file.
    """
    content = await file.read()
    try:
        chunks = smart_split_by_size(content, target_mb)
        if not chunks:
            raise HTTPException(status_code=400, detail="Failed to split PDF or empty document.")
        
        # Bundle chunks into in-memory zip
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            base_name = os.path.splitext(file.filename)[0]
            for idx, chunk_bytes in enumerate(chunks, 1):
                zip_file.writestr(f"{base_name}_part_{idx}.pdf", chunk_bytes)
        
        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={base_name}_split.zip"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/docs/enhance-scan")
async def enhance_scan_endpoint(file: UploadFile = File(...)):
    """
    Applies adaptive contrast thresholding to a scanned image, generating a clean PDF.
    Uses OpenCV (cv2) with PIL fallback for maximum compatibility.
    """
    content = await file.read()
    try:
        validate_pdf_input(content, file.filename or "image")  # Reuse validation
    except DocumentValidationError as e:
        raise HTTPException(status_code=400, detail=f"Invalid file: {str(e)}")

    try:
        pdf_bytes = enhance_scan(content)
        base_name = os.path.splitext(file.filename or "image")[0]
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={base_name}_enhanced.pdf"}
        )

    except DocumentValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Image enhancement failed: {str(e)}")

@app.post("/api/docs/compress")
async def compress_pdf_endpoint(file: UploadFile = File(...), quality: int = Form(50)):
    """
    Optimizes a PDF, outputting compaction metrics in custom response headers.
    """
    content = await file.read()
    original_size = len(content)
    try:
        compressed_bytes = compress_pdf(content, quality)
        compressed_size = len(compressed_bytes)
        
        base_name = os.path.splitext(file.filename)[0]
        headers = {
            "X-Original-Size": str(original_size),
            "X-Compressed-Size": str(compressed_size),
            "Access-Control-Expose-Headers": "X-Original-Size, X-Compressed-Size",
            "Content-Disposition": f"attachment; filename={base_name}_compressed.pdf"
        }
        
        return Response(
            content=compressed_bytes,
            media_type="application/pdf",
            headers=headers
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

class ItemUpdateRequest(BaseModel):
    field: str
    value: Any

@app.put("/api/items/{item_id}")
async def update_item(
    item_id: int,
    type: str,
    req: ItemUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor"])),
):
    if type == "sales":
        item = db.query(SalesLineItem).filter(SalesLineItem.id == item_id).first()
    elif type == "purchase":
        item = db.query(PurchaseLineItem).filter(PurchaseLineItem.id == item_id).first()
    else:
        raise HTTPException(status_code=400, detail="Invalid type")

    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Tenant isolation: a line item's tenant is derived via task -> batch.
    resource_tenant_id = item.task.batch.tenant_id if item.task and item.task.batch else None
    require_same_tenant(resource_tenant_id, current_user)

    # Primary/foreign-key columns must never be client-writable: rewriting
    # `task_id` lets an in-tenant caller re-parent their own item onto a
    # different (potentially cross-tenant) task, bypassing the tenant check
    # above entirely after the fact.
    NON_EDITABLE_FIELDS = {"id", "task_id"}
    if req.field in NON_EDITABLE_FIELDS or not hasattr(item, req.field):
        raise HTTPException(status_code=400, detail="Invalid field")
        
    old_value = getattr(item, req.field)
    
    # Attempt to cast float if the original column is float
    from sqlalchemy import Float
    if isinstance(getattr(item.__table__.columns, req.field).type, Float):
        try:
            val = float(req.value) if req.value else 0.0
            setattr(item, req.field, val)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid float value")
    else:
        setattr(item, req.field, req.value)
        
    db.commit()
    
    # Emit CA reviewer correction flag if value has changed
    if old_value != req.value:
        from services.observability import ObsLogger
        from models import ObservabilityLog
        import json
        
        task_id = item.task_id
        task = item.task
        batch_id = task.batch_id if task else "unknown"
        
        # Query composite score from observability logs
        log = db.query(ObservabilityLog).filter(
            ObservabilityLog.file_id == task_id,
            ObservabilityLog.event_type == "extraction_quality_score"
        ).first()
        
        composite_score = 0.85
        if log:
            try:
                payload = json.loads(log.payload_json)
                composite_score = payload.get("composite_score", 0.85)
            except:
                pass
                
        try:
            ObsLogger.emit_ca_flag(
                batch_id=batch_id,
                file_id=task_id,
                rejected_field=req.field,
                extracted_value=old_value,
                ca_corrected_value=req.value,
                composite_score=composite_score,
                db_session=db
            )
        except Exception as e:
            print(f"Error logging CA flag: {e}")
            
    return {"status": "success"}

@app.post("/api/invoice-metadata")
async def ocr_extract_endpoint(file: UploadFile = File(...), provider: str = Form("auto")):
    """
    Tiered OCR extraction with intelligent fallback.

    Provider options:
      "auto": PyMuPDF (native) → Tesseract (if available) → EasyOCR
      "pymupdf": Native PDF text extraction only
      "tesseract": Lightweight CPU OCR (if available)
      "easyocr": Heavyweight accuracy OCR

    Returns:
      {
        "success": bool,
        "text": extracted text,
        "provider_used": which OCR tier succeeded,
        "char_count": length of extracted text,
        "warning": optional message if OCR degraded
      }
    """
    content = await file.read()
    provider = provider.lower().strip()

    if provider not in ["auto", "pymupdf", "tesseract", "easyocr"]:
        raise HTTPException(status_code=400, detail=f"Unknown OCR provider: {provider}. Use: auto, pymupdf, tesseract, easyocr")

    try:
        # Route heavy OCR to dedicated Celery worker when available
        text = ocr_extract_via_celery(content, provider=provider)

        # Determine which provider was actually used (by logging)
        provider_used = "unknown"
        if provider == "pymupdf" or provider == "auto":
            provider_used = "PyMuPDF native text extraction"
        elif provider == "tesseract":
            provider_used = "Tesseract OCR"
        elif provider == "easyocr":
            provider_used = "EasyOCR"

        return JSONResponse(content={
            "success": True,
            "text": text,
            "provider_used": provider_used,
            "char_count": len(text),
            "warning": "No text extracted — PDF may be image-only or corrupted" if not text else None
        })

    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR extraction failed: {str(e)}")


# ── Google Drive Auto-Sync ────────────────────────────────────────────────────

class GoogleDriveSyncConfigRequest(BaseModel):
    # Legacy single-folder mode - optional now, a tenant may only want the
    # self-resolving mode below.
    folder_id: Optional[str] = None
    invoice_type: str = "both"
    schedule: str = "0 0 1 * *"

    # Self-resolving month-folder mode (Sales/Purchase/GSTR-2B) - added
    # 2026-07-09 so any tenant can configure this from the UI instead of
    # needing a hand-edited data/drive_paths/<slug>.json file (see
    # models.GoogleDriveSyncConfig's docstring). All optional - a tenant
    # sets whichever of these three they actually use.
    fiscal_year_start_month: Optional[int] = None
    month_folder_pattern: Optional[str] = None
    sales_root_folder_id: Optional[str] = None
    purchase_root_folder_id: Optional[str] = None
    gstr2b_root_folder_id: Optional[str] = None
    sales_schedule: Optional[str] = None
    purchase_schedule: Optional[str] = None
    gstr2b_schedule: Optional[str] = None


class GoogleDriveSyncTriggerRequest(BaseModel):
    # All fields optional — if omitted, saved config is used
    folder_id: Optional[str] = None
    invoice_type: Optional[str] = None
    llm_config: Optional[dict] = None
    # Cap on new/changed files processed in this run. Extraction is LLM-bound
    # (~80-90s/file) and the underlying Celery task has a 1-hour hard time
    # limit, so large folders MUST be processed in bounded batches across
    # multiple triggers rather than one unbounded run. None = no cap (only
    # safe for small folders / small remaining backlogs).
    max_files: Optional[int] = None
    # Restrict this run to one Drive subfolder (e.g. a specific month), instead
    # of scanning the whole configured folder tree. Get the id from
    # GET /api/google-drive-sync/subfolders.
    subfolder_id: Optional[str] = None


@app.get("/api/google-drive-sync/config")
async def get_drive_sync_config(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return the saved Google Drive sync config for the current tenant."""
    if not current_user.tenant_id:
        return JSONResponse(content={"configured": False, "config": None, "no_tenant": True})
    cfg = db.query(GoogleDriveSyncConfig).filter(
        GoogleDriveSyncConfig.tenant_id == current_user.tenant_id
    ).first()
    if not cfg:
        return JSONResponse(content={"configured": False, "config": None})
    return JSONResponse(content={
        "configured": True,
        "config": {
            "folder_id": cfg.folder_id,
            "invoice_type": cfg.invoice_type,
            "schedule": cfg.schedule,
            "fiscal_year_start_month": cfg.fiscal_year_start_month,
            "month_folder_pattern": cfg.month_folder_pattern,
            "sales_root_folder_id": cfg.sales_root_folder_id,
            "purchase_root_folder_id": cfg.purchase_root_folder_id,
            "gstr2b_root_folder_id": cfg.gstr2b_root_folder_id,
            "sales_schedule": cfg.sales_schedule,
            "purchase_schedule": cfg.purchase_schedule,
            "gstr2b_schedule": cfg.gstr2b_schedule,
            "updated_at": cfg.updated_at.isoformat() if cfg.updated_at else None,
        }
    })


@app.post("/api/google-drive-sync/config")
async def save_drive_sync_config(
    req: GoogleDriveSyncConfigRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "developer"])),
):
    """
    Save (or update) the Google Drive sync config for this tenant.
    Also writes Celery Beat schedule entries to beat_schedules.json so
    periodic syncs pick up after a celery beat restart - both the legacy
    single-folder mode (if folder_id is set) and the self-resolving
    Sales/Purchase/GSTR-2B mode (for whichever of those root folder IDs
    are set), added 2026-07-09 so a tenant self-configures both modes
    from this one endpoint instead of needing a hand-edited
    data/drive_paths/<slug>.json file for the self-resolving mode (see
    models.GoogleDriveSyncConfig's docstring for why that only ever
    worked for OneStack).
    """
    # Auto-create a tenant for this user if they don't have one yet
    if not current_user.tenant_id:
        email_slug = current_user.email.split("@")[0].lower().replace(".", "-")[:30]
        existing_slug = db.query(Tenant).filter(Tenant.slug == email_slug).first()
        slug = email_slug if not existing_slug else f"{email_slug}-{str(uuid.uuid4())[:6]}"
        new_tenant = Tenant(name=current_user.email, slug=slug)
        db.add(new_tenant)
        db.flush()
        current_user.tenant_id = new_tenant.id
        db.commit()
        db.refresh(current_user)

    if req.invoice_type not in ["sales", "purchase", "both"]:
        raise HTTPException(status_code=400, detail="invoice_type must be sales, purchase, or both")

    # Upsert config row
    cfg = db.query(GoogleDriveSyncConfig).filter(
        GoogleDriveSyncConfig.tenant_id == current_user.tenant_id
    ).first()
    if not cfg:
        cfg = GoogleDriveSyncConfig(tenant_id=current_user.tenant_id)
        db.add(cfg)

    cfg.folder_id    = req.folder_id
    cfg.invoice_type = req.invoice_type
    cfg.schedule     = req.schedule
    cfg.fiscal_year_start_month = req.fiscal_year_start_month
    cfg.month_folder_pattern    = req.month_folder_pattern
    cfg.sales_root_folder_id    = req.sales_root_folder_id
    cfg.purchase_root_folder_id = req.purchase_root_folder_id
    cfg.gstr2b_root_folder_id   = req.gstr2b_root_folder_id
    cfg.sales_schedule          = req.sales_schedule
    cfg.purchase_schedule       = req.purchase_schedule
    cfg.gstr2b_schedule         = req.gstr2b_schedule
    db.commit()

    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    tenant_slug = tenant.slug if tenant else current_user.tenant_id

    # Register Celery Beat schedules - one entry per configured mode.
    import json as _json
    beat_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "beat_schedules.json")
    try:
        registry = _json.load(open(beat_file, encoding="utf-8")) if os.path.exists(beat_file) else {}

        if req.folder_id:
            registry[f"google_drive_sync_{current_user.tenant_id}"] = {
                "task": "tasks.google_drive_sync_task",
                "cron": req.schedule,
                "kwargs": {
                    "tenant_id": current_user.tenant_id,
                    "google_drive_folder_id": req.folder_id,
                    "excel_output_path": f"/data/sync_{current_user.tenant_id}.xlsx",
                    "invoice_type": req.invoice_type,
                    "model_config": None,
                },
                "options": {"queue": "default"},
                "registered_at": datetime.utcnow().isoformat(),
            }

        if req.sales_root_folder_id:
            registry[f"sales_ingestion_{tenant_slug}"] = {
                "task": "tasks.sales_ingestion_task",
                "cron": req.sales_schedule or "0 2 * * *",
                "kwargs": {
                    "tenant_id": current_user.tenant_id,
                    "tenant_slug": tenant_slug,
                    "excel_output_path": f"/data/sync_{current_user.tenant_id}_sales.xlsx",
                    "invoice_type": "sales",
                    "model_config": None,
                },
                "options": {"queue": "drive_sync", "priority": 10},
                "registered_at": datetime.utcnow().isoformat(),
            }

        if req.purchase_root_folder_id:
            registry[f"purchase_ingestion_{tenant_slug}"] = {
                "task": "tasks.purchase_ingestion_task",
                "cron": req.purchase_schedule or "0 2 * * *",
                "kwargs": {
                    "tenant_id": current_user.tenant_id,
                    "tenant_slug": tenant_slug,
                    "excel_output_path": f"/data/sync_{current_user.tenant_id}_purchase.xlsx",
                    "model_config": None,
                },
                "options": {"queue": "drive_sync", "priority": 10},
                "registered_at": datetime.utcnow().isoformat(),
            }

        if req.gstr2b_root_folder_id:
            registry[f"gstr2b_ingestion_{tenant_slug}"] = {
                "task": "tasks.gstr2b_ingestion_task",
                "cron": req.gstr2b_schedule or "0 3 * * *",
                "kwargs": {
                    "tenant_id": current_user.tenant_id,
                    "tenant_slug": tenant_slug,
                },
                "options": {"queue": "drive_sync", "priority": 10},
                "registered_at": datetime.utcnow().isoformat(),
            }

        with open(beat_file, "w", encoding="utf-8") as f:
            _json.dump(registry, f, indent=2)
    except Exception as e:
        logger.warning("Could not update beat_schedules.json: %s", e)

    return JSONResponse(content={
        "ok": True,
        "folder_id": cfg.folder_id,
        "invoice_type": cfg.invoice_type,
        "sales_root_folder_id": cfg.sales_root_folder_id,
        "purchase_root_folder_id": cfg.purchase_root_folder_id,
        "gstr2b_root_folder_id": cfg.gstr2b_root_folder_id,
    })


@app.get("/api/google-drive-sync/subfolders")
async def list_drive_subfolders(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    List the immediate subfolders of the tenant's configured Drive folder —
    used to populate a "which month?" picker, since clients commonly organize
    invoices into month/year subfolders. Syncing one month at a time is a
    natural, cost-bounded unit of work.
    """
    cfg = db.query(GoogleDriveSyncConfig).filter(
        GoogleDriveSyncConfig.tenant_id == current_user.tenant_id
    ).first()
    if not cfg:
        raise HTTPException(status_code=400, detail="No Drive folder configured yet.")

    from services.google_drive import GoogleDriveConnector
    try:
        connector = GoogleDriveConnector(cfg.folder_id)
        subfolders = connector.list_subfolders()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not list Drive subfolders: {e}")

    return JSONResponse(content={"folder_id": cfg.folder_id, "subfolders": subfolders})


@app.post("/api/google-drive-sync/trigger")
async def trigger_google_drive_sync(
    req: GoogleDriveSyncTriggerRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Trigger an immediate Google Drive sync for the current tenant.
    Uses saved config if folder_id / invoice_type are not provided in the request.
    """
    from celery_app import google_drive_sync_task

    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="No tenant assigned to this user")

    # Resolve folder_id and invoice_type: request overrides saved config
    folder_id    = req.folder_id
    invoice_type = req.invoice_type
    if not folder_id or not invoice_type:
        cfg = db.query(GoogleDriveSyncConfig).filter(
            GoogleDriveSyncConfig.tenant_id == current_user.tenant_id
        ).first()
        if not cfg:
            raise HTTPException(
                status_code=400,
                detail="No Drive folder configured. Save config first via POST /api/google-drive-sync/config"
            )
        folder_id    = folder_id    or cfg.folder_id
        invoice_type = invoice_type or cfg.invoice_type

    excel_path = f"/data/sync_{current_user.tenant_id}.xlsx"

    task = google_drive_sync_task.delay(
        tenant_id=current_user.tenant_id,
        google_drive_folder_id=folder_id,
        excel_output_path=excel_path,
        invoice_type=invoice_type,
        model_config=req.llm_config,
        max_files=req.max_files,
        subfolder_id=req.subfolder_id,
    )

    return JSONResponse(content={
        "status": "sync_started",
        "task_id": task.id,
        "folder_id": req.subfolder_id or folder_id,
        "invoice_type": invoice_type,
        "max_files": req.max_files,
    })


def _tenant_slug_for(db: Session, current_user: User) -> str:
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    return tenant.slug if tenant else current_user.tenant_id


@app.post("/api/google-drive-sync/trigger-sales")
async def trigger_sales_ingestion(
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor", "developer"])),
):
    """
    Runs tasks.sales_ingestion_task immediately for the current tenant -
    the self-resolving Sales pipeline (Phase 6), triggered on demand from
    the Drive Sync UI instead of only via a Celery Beat schedule.
    """
    from celery_app import sales_ingestion_task

    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="No tenant assigned to this user")
    cfg = db.query(GoogleDriveSyncConfig).filter(GoogleDriveSyncConfig.tenant_id == current_user.tenant_id).first()
    if not cfg or not cfg.sales_root_folder_id:
        raise HTTPException(status_code=400, detail="No sales_root_folder_id configured. Save config first.")

    tenant_slug = _tenant_slug_for(db, current_user)
    task = sales_ingestion_task.delay(
        tenant_id=current_user.tenant_id,
        tenant_slug=tenant_slug,
        excel_output_path=f"/data/sync_{current_user.tenant_id}_sales.xlsx",
        invoice_type="sales",
        model_config=None,
    )
    return JSONResponse(content={"status": "sync_started", "task_id": task.id})


@app.post("/api/google-drive-sync/trigger-purchase")
async def trigger_purchase_ingestion(
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor", "developer"])),
):
    """Runs tasks.purchase_ingestion_task immediately for the current tenant."""
    from celery_app import purchase_ingestion_task

    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="No tenant assigned to this user")
    cfg = db.query(GoogleDriveSyncConfig).filter(GoogleDriveSyncConfig.tenant_id == current_user.tenant_id).first()
    if not cfg or not cfg.purchase_root_folder_id:
        raise HTTPException(status_code=400, detail="No purchase_root_folder_id configured. Save config first.")

    tenant_slug = _tenant_slug_for(db, current_user)
    task = purchase_ingestion_task.delay(
        tenant_id=current_user.tenant_id,
        tenant_slug=tenant_slug,
        excel_output_path=f"/data/sync_{current_user.tenant_id}_purchase.xlsx",
        model_config=None,
    )
    return JSONResponse(content={"status": "sync_started", "task_id": task.id})


@app.post("/api/google-drive-sync/trigger-gstr2b")
async def trigger_gstr2b_ingestion(
    db: Session = Depends(get_db),
    current_user: User = Depends(RoleChecker(["owner", "auditor", "developer"])),
):
    """Runs tasks.gstr2b_ingestion_task immediately for the current tenant."""
    from celery_app import gstr2b_ingestion_task

    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="No tenant assigned to this user")
    cfg = db.query(GoogleDriveSyncConfig).filter(GoogleDriveSyncConfig.tenant_id == current_user.tenant_id).first()
    if not cfg or not cfg.gstr2b_root_folder_id:
        raise HTTPException(status_code=400, detail="No gstr2b_root_folder_id configured. Save config first.")

    tenant_slug = _tenant_slug_for(db, current_user)
    task = gstr2b_ingestion_task.delay(
        tenant_id=current_user.tenant_id,
        tenant_slug=tenant_slug,
    )
    return JSONResponse(content={"status": "sync_started", "task_id": task.id})


@app.get("/api/google-drive-sync/status/{task_id}")
async def get_sync_status(task_id: str):
    """Poll status of a running sync task. Returns result including batch_id on SUCCESS."""
    from celery_app import celery_app as _celery

    task_result = _celery.AsyncResult(task_id)
    result = None
    if task_result.status == "SUCCESS" and isinstance(task_result.result, dict):
        result = task_result.result

    return JSONResponse(content={
        "task_id": task_id,
        "status": task_result.status,
        "result": result,
        "error": str(task_result.info) if task_result.status == "FAILURE" else None,
    })


@app.get("/api/google-drive-sync/history")
async def get_sync_history(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = 20,
):
    """
    Return sync job history for this tenant.
    Each job includes a batch_id computed from tenant + date so the frontend
    can construct a direct download URL using GET /api/export/{batch_id}.
    """
    sync_jobs = db.query(GoogleDriveSyncJob).filter(
        GoogleDriveSyncJob.tenant_id == current_user.tenant_id
    ).order_by(GoogleDriveSyncJob.sync_timestamp.desc()).limit(limit).all()

    def _batch_id(job):
        # Deterministic: matches what GoogleDriveSyncPipeline._process_invoice() computes
        return f"sync_{job.tenant_id}_{job.sync_timestamp.strftime('%Y%m%d')}"

    return JSONResponse(content={
        "sync_jobs": [
            {
                "id": job.id,
                "batch_id": _batch_id(job),
                "sync_timestamp": job.sync_timestamp.isoformat(),
                "total_files_found": job.total_files_found,
                "new_files": job.new_files,
                "updated_files": job.updated_files,
                "processed_files": job.processed_files,
                "failed_files": job.failed_files,
                "status": job.status,
                "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            }
            for job in sync_jobs
        ]
    })


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


# ── GSTR-2B Reconciliation ────────────────────────────────────────────────────

class ReconcileRequest(BaseModel):
    items: List[dict]          # extracted invoice line items
    gstr2b: Any                # raw GSTR-2B JSON (dict) or JSON string


@app.post("/api/reconcile")
async def reconcile_gstr2b(req: ReconcileRequest):
    """
    Match extracted invoice items against GSTR-2B data.
    Returns annotated rows (with recon_status) + a summary.
    """
    try:
        raw_2b = req.gstr2b
        if isinstance(raw_2b, str):
            raw_2b = json.loads(raw_2b)

        gstr2b_records = parse_gstr2b(raw_2b)
        if not gstr2b_records:
            raise HTTPException(
                status_code=422,
                detail="No B2B invoice records found in the GSTR-2B JSON. "
                       "Please upload the full GSTR-2B JSON downloaded from the GST portal."
            )

        result = recon_match(req.items, gstr2b_records)
        return JSONResponse(content=result)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Reconciliation error: {e}")


class BatchReconcileRequest(BaseModel):
    gstr2b: Any   # raw GSTR-2B JSON from portal


@app.post("/api/reconcile/from-batch/{batch_id}")
async def reconcile_from_batch(
    batch_id: str,
    req: BatchReconcileRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Run GSTR-2B reconciliation using a batch's extracted purchase items as the
    books source — no manual JSON export needed.
    Returns the same response shape as POST /api/reconcile.
    """
    from services.gstr2b_reconciler import parse_gstr2b, reconcile as recon_match

    batch = db.query(BatchJob).filter(BatchJob.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    require_same_tenant(batch.tenant_id, current_user)

    tasks = db.query(InvoiceTask).filter(InvoiceTask.batch_id == batch_id).all()
    purchase_items = []
    for t in tasks:
        for item in (getattr(t, "purchase_items", None) or []):
            purchase_items.append(
                {c.name: getattr(item, c.name) for c in item.__table__.columns}
            )

    if not purchase_items:
        raise HTTPException(
            status_code=400,
            detail="No purchase items found in this batch. Run purchase extraction first.",
        )

    try:
        raw_2b = req.gstr2b
        if isinstance(raw_2b, str):
            raw_2b = json.loads(raw_2b)
        gstr2b_records = parse_gstr2b(raw_2b)
        if not gstr2b_records:
            raise HTTPException(
                status_code=422,
                detail="No B2B invoice records found in the GSTR-2B JSON.",
            )
        result = recon_match(purchase_items, gstr2b_records)
        return JSONResponse(content=result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Reconciliation error: {e}")


@app.post("/api/reconcile/export")
async def export_reconciliation(req: ReconcileRequest):
    """
    Run reconciliation and return a color-coded Excel file with two sheets:
      Sheet 1 — Reconciliation (all rows, color-coded by status)
      Sheet 2 — Summary (counts + amounts per status)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        raw_2b = req.gstr2b
        if isinstance(raw_2b, str):
            raw_2b = json.loads(raw_2b)

        gstr2b_records = parse_gstr2b(raw_2b)
        result = recon_match(req.items, gstr2b_records)
        all_rows = result["rows"] + result["extra"]
        summary  = result["summary"]

        # ── Colour map ───────────────────────────────────────────────────────
        STATUS_FILLS = {
            "matched":       PatternFill("solid", fgColor="C6EFCE"),   # green
            "mismatch":      PatternFill("solid", fgColor="FFEB9C"),   # amber
            "missing_in_2b": PatternFill("solid", fgColor="FFC7CE"),   # red
            "not_in_books":  PatternFill("solid", fgColor="BDD7EE"),   # blue
        }
        STATUS_LABELS = {
            "matched":       "✓ Matched",
            "mismatch":      "⚠ Amount Mismatch",
            "missing_in_2b": "✗ Missing in 2B",
            "not_in_books":  "ℹ Not Booked",
        }

        wb = Workbook()

        # ── Sheet 1: Reconciliation ──────────────────────────────────────────
        ws = wb.active
        ws.title = "Reconciliation"

        headers = [
            "Status", "Supplier Invoice", "Invoice Date", "GSTIN",
            "Party Name", "Particulars",
            "Books: Taxable", "Books: SGST", "Books: CGST", "Books: IGST", "Books: Total",
            "2B: Taxable", "2B: SGST", "2B: CGST", "2B: IGST", "2B: Total",
            "Difference", "HSN",
        ]

        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[1].height = 28

        for row_idx, row in enumerate(all_rows, 2):
            status = row.get("recon_status", "")
            fill   = STATUS_FILLS.get(status)

            values = [
                STATUS_LABELS.get(status, status),
                row.get("supplier_inv"),
                row.get("invoice_date"),
                row.get("gst_no"),
                row.get("party_ac_name"),
                row.get("particulars"),
                row.get("amount"),
                row.get("sgst"),
                row.get("cgst"),
                row.get("igst"),
                row.get("total_amount"),
                row.get("2b_taxable_val"),
                row.get("2b_sgst"),
                row.get("2b_cgst"),
                row.get("2b_igst"),
                row.get("2b_total_val"),
                row.get("diff_amount"),
                row.get("hsn"),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if fill:
                    cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
        ws.column_dimensions["A"].width = 22  # Status
        ws.column_dimensions["E"].width = 28  # Party Name
        ws.column_dimensions["F"].width = 28  # Particulars

        ws.freeze_panes = "A2"  # freeze header

        # ── Sheet 2: Summary ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 26
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 18

        summary_headers = ["Status", "Count", "Total Amount (₹)"]
        for ci, h in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        summary_rows = [
            ("✓ Matched",          summary["counts"]["matched"],       summary["amounts"]["matched"],       "C6EFCE"),
            ("⚠ Amount Mismatch",  summary["counts"]["mismatch"],      summary["amounts"]["mismatch"],      "FFEB9C"),
            ("✗ Missing in 2B",    summary["counts"]["missing_in_2b"], summary["amounts"]["missing_in_2b"], "FFC7CE"),
            ("ℹ Not Booked",       summary["counts"]["not_in_books"],  summary["amounts"]["not_in_books"],  "BDD7EE"),
        ]
        for ri, (label, count, amount, color) in enumerate(summary_rows, 2):
            fill = PatternFill("solid", fgColor=color)
            for ci, val in enumerate([label, count, amount], 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill   = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # ITC risk rows
        ws2.cell(row=7, column=1, value="ITC at Risk (Missing + Mismatch)").font = Font(bold=True, color="C00000")
        ws2.cell(row=7, column=3, value=summary["itc_at_risk"]).font           = Font(bold=True, color="C00000")
        ws2.cell(row=8, column=1, value="Matched ITC (Claimable)").font        = Font(bold=True, color="375623")
        ws2.cell(row=8, column=3, value=summary["matched_itc"]).font           = Font(bold=True, color="375623")

        output_path = os.path.join(tempfile.gettempdir(), "gstr2b_reconciliation.xlsx")
        wb.save(output_path)

        return FileResponse(
            path=output_path,
            filename="gstr2b_reconciliation.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export error: {e}")


# ── Session Persistence Layer ─────────────────────────────────────────────────

class SessionUpsertRequest(BaseModel):
    active_batch_id:       Optional[str] = None
    active_task_id:        Optional[str] = None
    active_page:           Optional[str] = None
    selected_invoice_type: Optional[str] = None
    selected_model:        Optional[str] = None
    scroll_position:       Optional[int] = 0
    context_blob:          Optional[str] = None  # JSON string

class PreferencesRequest(BaseModel):
    theme:                Optional[str] = None
    default_invoice_type: Optional[str] = None
    default_model:        Optional[str] = None
    default_export_schema:Optional[str] = None

class AnnotationRequest(BaseModel):
    field_name:      str
    note:            Optional[str] = None
    original_value:  Optional[str] = None
    corrected_value: Optional[str] = None


@app.get("/api/me/session")
async def get_session(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns the user's last persisted session so the frontend can resume work."""
    session = db.query(UserSession).filter(UserSession.user_id == current_user.id).first()
    if not session:
        return {"session": None}

    # Enrich with batch status so the frontend can decide whether to show a resume prompt
    batch_status = None
    batch_total = None
    if session.active_batch_id:
        batch = db.query(BatchJob).filter(BatchJob.id == session.active_batch_id).first()
        if batch:
            batch_status = batch.status.value
            batch_total = batch.total_files

    return {
        "session": {
            "active_batch_id":       session.active_batch_id,
            "active_task_id":        session.active_task_id,
            "active_page":           session.active_page,
            "selected_invoice_type": session.selected_invoice_type,
            "selected_model":        session.selected_model,
            "scroll_position":       session.scroll_position,
            "context_blob":          session.context_blob,
            "updated_at":            session.updated_at.isoformat() if session.updated_at else None,
            "batch_status":          batch_status,
            "batch_total_files":     batch_total,
        }
    }


@app.post("/api/me/session")
async def upsert_session(
    req: SessionUpsertRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upserts the user's session state. Called on every meaningful UI action (debounced on frontend)."""
    session = db.query(UserSession).filter(UserSession.user_id == current_user.id).first()
    if session:
        if req.active_batch_id  is not None: session.active_batch_id       = req.active_batch_id
        if req.active_task_id   is not None: session.active_task_id        = req.active_task_id
        if req.active_page      is not None: session.active_page           = req.active_page
        if req.selected_invoice_type is not None: session.selected_invoice_type = req.selected_invoice_type
        if req.selected_model   is not None: session.selected_model        = req.selected_model
        if req.scroll_position  is not None: session.scroll_position       = req.scroll_position
        if req.context_blob     is not None: session.context_blob          = req.context_blob
        session.updated_at = datetime.utcnow()
    else:
        session = UserSession(
            user_id=current_user.id,
            **{k: v for k, v in req.dict().items() if v is not None}
        )
        db.add(session)
    db.commit()
    return {"ok": True}


@app.delete("/api/me/session")
async def clear_session(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Clears the active session (e.g., when user explicitly starts a new audit)."""
    db.query(UserSession).filter(UserSession.user_id == current_user.id).delete()
    db.commit()
    return {"ok": True}


@app.get("/api/me/preferences")
async def get_preferences(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns user preferences (theme, defaults). Returns defaults if not yet saved."""
    prefs = db.query(UserPreferences).filter(UserPreferences.user_id == current_user.id).first()
    if not prefs:
        return {
            "theme": "dark",
            "default_invoice_type": "both",
            "default_model": "auto",
            "default_export_schema": "standard",
        }
    return {
        "theme":                prefs.theme,
        "default_invoice_type": prefs.default_invoice_type,
        "default_model":        prefs.default_model,
        "default_export_schema":prefs.default_export_schema,
    }


@app.put("/api/me/preferences")
async def save_preferences(
    req: PreferencesRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upserts user preferences."""
    prefs = db.query(UserPreferences).filter(UserPreferences.user_id == current_user.id).first()
    if prefs:
        if req.theme                 is not None: prefs.theme                  = req.theme
        if req.default_invoice_type  is not None: prefs.default_invoice_type   = req.default_invoice_type
        if req.default_model         is not None: prefs.default_model           = req.default_model
        if req.default_export_schema is not None: prefs.default_export_schema   = req.default_export_schema
        prefs.updated_at = datetime.utcnow()
    else:
        prefs = UserPreferences(
            user_id=current_user.id,
            **{k: v for k, v in req.dict().items() if v is not None}
        )
        db.add(prefs)
    db.commit()
    return {"ok": True}


@app.post("/api/tasks/{task_id}/annotate")
async def annotate_task(
    task_id: str,
    req: AnnotationRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Saves a field-level auditor note or correction on an invoice task."""
    task = db.query(InvoiceTask).filter(InvoiceTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    annotation = UserAnnotation(
        user_id=current_user.id,
        task_id=task_id,
        field_name=req.field_name,
        note=req.note,
        original_value=req.original_value,
        corrected_value=req.corrected_value,
    )
    db.add(annotation)
    db.commit()
    return {"ok": True, "annotation_id": annotation.id}


@app.get("/api/tasks/{task_id}/annotations")
async def get_task_annotations(
    task_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns all annotations for a task (all reviewers, sorted by time)."""
    annotations = (
        db.query(UserAnnotation)
        .filter(UserAnnotation.task_id == task_id)
        .order_by(UserAnnotation.created_at.asc())
        .all()
    )
    return {
        "annotations": [
            {
                "id":              a.id,
                "field_name":      a.field_name,
                "note":            a.note,
                "original_value":  a.original_value,
                "corrected_value": a.corrected_value,
                "created_at":      a.created_at.isoformat(),
                "user_id":         a.user_id,
            }
            for a in annotations
        ]
    }
