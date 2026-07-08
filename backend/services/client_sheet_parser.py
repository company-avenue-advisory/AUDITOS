"""
Parses OneStack's client-supplied Excel sheet (classified as CLIENT_SHEET by
drive_classifier.py - any .xlsx sitting loose at the month-folder root, e.g.
"Sales_PQB_June 2026_Vandana.xlsx").

Real structure confirmed against the actual June 2026 file this session:
  - Sheet name contains "masterdata" (real example: "Sales_masterdata(Input)")
  - Row 1 is column-type annotations ("Input"/"Formula based"/HSN codes),
    not data - always skip it.
  - Row 2 is the real header row.
  - Data starts row 3. Both invoices AND credit notes live in this ONE
    sheet, distinguished by the "Document Type Code" column - NOT in the
    separately-named "Sales_Creditnote_master" tab, which turned out to be
    a stale prior-period batch unrelated to the current month (confirmed:
    that tab held December 2025 credit notes while the June 2026 file's
    real June credit notes were inside Sales_masterdata(Input) all along).

This module only parses and returns rows in a normalized shape - it does
NOT judge whether the client's figures are correct. That judgment belongs
to sales_reconciliation.py, which compares this parser's output against
our own extracted Sales Register.

Root cause identified for the client's systematic tax-type errors (see
sales_reconciliation.py's docstring for the full finding): the client's
"Interstate or Intrastate (Drop Down only)" column is confirmed via
openpyxl formula inspection to be a plain manually-typed string, not a
spreadsheet formula. Cross-checking it against "State of Supply" (which
IS a formula, "=LEFT(recipient_gstin, 2)") shows 76 of 196 June 2026
rows (39%) are the EXACT logical inversion of what the GSTIN-derived
state comparison implies - zero exceptions, not random entry error. This
parser returns both "supplier_location" and "state_of_supply" (in
addition to the client's own "interstate_or_intrastate" label and actual
entered tax amounts) specifically so sales_reconciliation.py can name
this known pattern when it recurs, rather than reporting a generic
"tax-type contradiction" every time.
"""
from typing import List, Optional

# real header text -> normalized key. Matched by substring (case/whitespace
# -insensitive) since these headers contain embedded newlines and irregular
# spacing in the real file - a header COULD shift by a column if the client
# reorders things, so this module resolves by NAME, never by column index.
_HEADER_MAP = {
    "s.no": "sno",
    "document type code": "doc_type",
    "document  no": "doc_no",
    "document no": "doc_no",
    "document  date": "doc_date",
    "document date": "doc_date",
    "supplier location": "supplier_location",
    "recipient billing name": "party_name",
    "recipient billing gstin": "party_gstin",
    "b2b or b2c": "b2b_or_b2c",
    "state  of supply": "state_of_supply",
    "state of supply": "state_of_supply",
    "interstate or intrastate": "interstate_or_intrastate",
    "net basic amt": "taxable",
    "invoice value": "total",
}

# raw column header (normalized, whitespace-collapsed) -> section key, for
# the per-section Gross-Discount-Advance breakdown. Only the NET
# ("Gross- Discount-Advance") column per section is kept - Gross/Discount/
# Advance individually are useful for audit trail but not needed for the
# headline taxable-value cross-check.
_SECTION_NET_HEADERS = {
    "saas (gross- discount-advance)": "saas_net",
    "transactional charges (gross- discount-advance)": "transactional_net",
    "kyc charges (gross- discount-advance)": "kyc_net",
    "promotional chrs (gross- discount-advance)": "promotional_net",
    "late charges": "late_charges",
    "sound (gross- discount-advance)": "soundbox_net",
}


def _normalize_header(h) -> str:
    if h is None:
        return ""
    return " ".join(str(h).replace("\n", " ").split()).lower()


def _find_masterdata_sheet(wb) -> Optional[str]:
    for name in wb.sheetnames:
        if "masterdata" in name.lower():
            return name
    return None


def parse_client_sheet(xlsx_path: str) -> List[dict]:
    """
    Returns one dict per row (invoice or credit note) from the client's
    masterdata sheet. Each dict has:
      sno, doc_type ("Invoice" | "Credit Note" | "Debit Note"), doc_no,
      doc_date, party_name, party_gstin, b2b_or_b2c, state_of_supply,
      interstate_or_intrastate, taxable, igst, sgst, cgst, total,
      sections: {saas_net, transactional_net, kyc_net, promotional_net,
                 late_charges, soundbox_net}

    Raises ValueError if no sheet with "masterdata" in its name is found -
    a client sheet that doesn't follow this convention needs a human to
    look at it, not a silent empty result.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = _find_masterdata_sheet(wb)
    if not sheet_name:
        raise ValueError(
            f"No sheet with 'masterdata' in its name found in {xlsx_path} "
            f"(sheets present: {wb.sheetnames}) - client sheet structure may have changed."
        )
    ws = wb[sheet_name]

    header_row = [_normalize_header(c.value) for c in ws[2]]
    col_map = {}  # normalized target key -> column index
    section_map = {}  # section key -> column index
    for i, h in enumerate(header_row):
        if not h:
            continue
        for prefix, key in _HEADER_MAP.items():
            if h.startswith(prefix) or h == prefix:
                col_map.setdefault(key, i)
        for prefix, key in _SECTION_NET_HEADERS.items():
            if h.startswith(prefix):
                section_map.setdefault(key, i)
        # exact-match IGST/SGST/CGST (avoid matching the numbered duplicate
        # columns further right, e.g. "IGST3", "IGST9", "IGST13", "IGST17")
        if h == "igst":
            col_map.setdefault("igst", i)
        if h == "sgst":
            col_map.setdefault("sgst", i)
        if h == "cgst":
            col_map.setdefault("cgst", i)

    def _get(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None else None

    def _get_section(row, key):
        idx = section_map.get(key)
        return row[idx] if idx is not None else None

    def _f(v):
        if v is None:
            return 0.0
        if isinstance(v, str):
            v = v.strip().replace(",", "")
            if not v or v == "-":
                return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def _date_str(v):
        if v is None:
            return None
        if hasattr(v, "strftime"):
            return v.strftime("%d-%m-%Y")
        return str(v)

    results = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        doc_no = _get(row, "doc_no")
        doc_type = _get(row, "doc_type")
        if not doc_no or not doc_type:
            continue
        results.append({
            "sno": _get(row, "sno"),
            "doc_type": doc_type,
            "doc_no": str(doc_no).strip(),
            "doc_date": _date_str(_get(row, "doc_date")),
            "party_name": _get(row, "party_name"),
            "party_gstin": _get(row, "party_gstin"),
            "b2b_or_b2c": _get(row, "b2b_or_b2c"),
            "supplier_location": _get(row, "supplier_location"),
            "state_of_supply": str(_get(row, "state_of_supply") or "").strip(),
            "interstate_or_intrastate": _get(row, "interstate_or_intrastate"),
            "taxable": round(_f(_get(row, "taxable")), 2),
            "igst": round(_f(_get(row, "igst")), 2),
            "sgst": round(_f(_get(row, "sgst")), 2),
            "cgst": round(_f(_get(row, "cgst")), 2),
            "total": round(_f(_get(row, "total")), 2),
            "sections": {
                "saas_net": round(_f(_get_section(row, "saas_net")), 2),
                "transactional_net": round(_f(_get_section(row, "transactional_net")), 2),
                "kyc_net": round(_f(_get_section(row, "kyc_net")), 2),
                "promotional_net": round(_f(_get_section(row, "promotional_net")), 2),
                "late_charges": round(_f(_get_section(row, "late_charges")), 2),
                "soundbox_net": round(_f(_get_section(row, "soundbox_net")), 2),
            },
        })
    return results
