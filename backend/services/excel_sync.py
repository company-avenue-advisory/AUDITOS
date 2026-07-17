"""
Excel sync service — append extraction results to existing workbook.

Handles:
  - Loading existing Excel file
  - Appending new rows with extracted invoice data
  - Maintaining structure and formatting
  - Tracking which row in Excel corresponds to which extraction task
"""

import os
import logging
from datetime import datetime
from typing import Optional, List
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Install it to enable Excel sync.")

logger = logging.getLogger(__name__)

# Canonical output schema — single source of truth for columns + value order.
# The register views below emit the exact same labels/order as before (format-
# preserving); see services/output_schema.py.
from services.output_schema import (
    SALES_REGISTER_VIEW,
    PURCHASE_REGISTER_VIEW,
    labels,
    row_for,
)


def _row_ctx(source_filename: str) -> dict:
    """Per-row context (source file + processed timestamp) for row_for()."""
    return {
        "processed_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": source_filename,
    }


class ExcelSyncService:
    """
    Manages appending extracted invoice data to Excel workbooks.
    One workbook per tenant, continuously appended with new invoices.
    """

    # Column labels are derived from the canonical views (unchanged output).
    SALES_COLUMNS = labels(SALES_REGISTER_VIEW)
    PURCHASE_COLUMNS = labels(PURCHASE_REGISTER_VIEW)

    def __init__(self, excel_path: str, invoice_type: str = "sales"):
        """
        Initialize Excel sync service.

        Args:
            excel_path: Path to Excel file (will be created if doesn't exist)
            invoice_type: "sales" or "purchase"
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed")

        self.excel_path = excel_path
        self.invoice_type = invoice_type.lower()
        if self.invoice_type not in ["sales", "purchase"]:
            raise ValueError("invoice_type must be 'sales' or 'purchase'")

        self._lock = None  # File lock (acquired per operation)

    def _get_workbook(self):
        """Load existing workbook or create new one. Acquires file lock."""
        from services.excel_lockfile import ExcelFileLock

        # Acquire lock to prevent concurrent writes
        self._lock = ExcelFileLock(self.excel_path, timeout=30)
        self._lock.acquire()

        if os.path.exists(self.excel_path):
            return load_workbook(self.excel_path)
        else:
            # Create new workbook with headers
            logger.info(f"Creating new Excel workbook at {self.excel_path}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"

            columns = self.SALES_COLUMNS if self.invoice_type == "sales" else self.PURCHASE_COLUMNS
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                # Format header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Auto-adjust column widths
            for col_idx, col_name in enumerate(columns, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15

            return wb

    def append_sales_item(self, item, source_filename: str) -> int:
        """
        Append a sales line item to Excel.

        Args:
            item: SalesLineItem ORM object with extracted data
            source_filename: Name of the invoice file processed

        Returns:
            Row number in Excel where item was appended
        """
        if self.invoice_type != "sales":
            raise ValueError("This method is for sales invoices only")

        wb = self._get_workbook()
        ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        row_data = row_for(SALES_REGISTER_VIEW, item, _row_ctx(source_filename))

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col_idx, value=value)

        self._safe_save(wb)
        logger.info(f"Appended sales item to {self.excel_path} at row {next_row}")
        return next_row

    def append_purchase_item(self, item, source_filename: str) -> int:
        """
        Append a purchase line item to Excel.

        Args:
            item: PurchaseLineItem ORM object
            source_filename: Name of the invoice file processed

        Returns:
            Row number in Excel where item was appended
        """
        if self.invoice_type != "purchase":
            raise ValueError("This method is for purchase invoices only")

        wb = self._get_workbook()
        ws = wb.active

        next_row = ws.max_row + 1

        row_data = row_for(PURCHASE_REGISTER_VIEW, item, _row_ctx(source_filename))

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col_idx, value=value)

        self._safe_save(wb)
        logger.info(f"Appended purchase item to {self.excel_path} at row {next_row}")
        return next_row

    def append_batch(self, items: List, source_filename: str, is_sales: bool = True) -> List[int]:
        """
        Append multiple items at once (more efficient).

        Args:
            items: List of SalesLineItem or PurchaseLineItem objects
            source_filename: Name of invoice file
            is_sales: True for sales, False for purchase

        Returns:
            List of row numbers where items were appended
        """
        if not items:
            return []

        wb = self._get_workbook()
        ws = wb.active

        row_ids = []
        next_row = ws.max_row + 1

        for item in items:
            if is_sales and self.invoice_type == "sales":
                row_data = row_for(SALES_REGISTER_VIEW, item, _row_ctx(source_filename))
            elif not is_sales and self.invoice_type == "purchase":
                row_data = row_for(PURCHASE_REGISTER_VIEW, item, _row_ctx(source_filename))
            else:
                continue

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col_idx, value=value)

            row_ids.append(next_row)
            next_row += 1

        self._safe_save(wb)
        logger.info(f"Appended {len(row_ids)} items to {self.excel_path}")
        return row_ids

    def get_max_row(self) -> int:
        """Get current maximum row in Excel (for tracking)."""
        if not os.path.exists(self.excel_path):
            return 1  # Header row only
        try:
            wb = load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            return ws.max_row
        except Exception as e:
            logger.error(f"Error getting max row: {e}")
            return 1

    def _release_lock(self):
        """Release file lock if held."""
        if self._lock:
            try:
                self._lock.release()
            except Exception as e:
                logger.warning(f"Error releasing lock: {e}")
            finally:
                self._lock = None

    def _safe_save(self, wb):
        """Save workbook and release lock."""
        try:
            wb.save(self.excel_path)
            logger.debug(f"Saved {self.excel_path}")
        finally:
            self._release_lock()
