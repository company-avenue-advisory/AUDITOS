"""
gstr2b_excel_export.py — Writes the GSTR-2B ↔ Purchase Register reconciliation
workbook (Section 13 of the AuditOS reconciliation workflow).

Input is the raw dict returned by services.gstr2b_reconciler.reconcile():
    {"rows": [...], "extra": [...], "summary": {...}}

"rows"  — books-sourced (matched / mismatch / missing_in_2b), one per line item.
"extra" — 2B-only (not_in_books): supplier filed it, nothing in our books.

Columns and labels come from services.output_schema.GSTR2B_RECONCILIATION_VIEW
(the canonical schema) — this file only handles the openpyxl mechanics and the
per-bucket color coding / gap-investigation columns (Section 13.1):

  not_in_books  -> "In Drive?" + "Entry Missing in Tally" filled by an optional
                   `drive_checker` callback (targeted Drive-verify step); when
                   Drive=Yes -> internal Tally-booking task, no client contact.
                   When Drive=No -> "Client Action" = request from client
                   (accountant-gated — see Section 13.1).
  missing_in_2b -> "Client Action" = auto vendor-followup notice (send itself
                   is a separate, policy-gated step — this only labels it).
  mismatch      -> "Client Action" = flag for accountant review (internal).
"""

import logging
from typing import Any, Callable, Optional

from services.output_schema import GSTR2B_RECONCILIATION_VIEW, labels, row_for

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Install it to enable GSTR-2B Excel export.")

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79") if OPENPYXL_AVAILABLE else None
HEADER_FONT = Font(bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
MATCHED_FILL       = PatternFill("solid", fgColor="E2EFDA") if OPENPYXL_AVAILABLE else None  # green
MISMATCH_FILL      = PatternFill("solid", fgColor="FFF2CC") if OPENPYXL_AVAILABLE else None  # amber
MISSING_IN_2B_FILL = PatternFill("solid", fgColor="FCE4D6") if OPENPYXL_AVAILABLE else None  # red
NOT_IN_BOOKS_FILL  = PatternFill("solid", fgColor="F8CBAD") if OPENPYXL_AVAILABLE else None  # orange
BORDER = Border(*(Side(style="thin"),) * 4) if OPENPYXL_AVAILABLE else None

_STATUS_FILL = {
    "matched":       MATCHED_FILL,
    "mismatch":      MISMATCH_FILL,
    "missing_in_2b": MISSING_IN_2B_FILL,
    "not_in_books":  NOT_IN_BOOKS_FILL,
}

_CLIENT_ACTION = {
    "mismatch":      "Flag for accountant review",
    "missing_in_2b": "Auto-notify client: follow up with vendor (ITC at risk)",
    # "not_in_books" client_action is set per-row below (depends on Drive result).
}


def _annotate_gap_row(row: dict, drive_checker: Optional[Callable[[dict], bool]]) -> dict:
    """Fills the Section 13.1 gap-investigation columns onto a `not_in_books` row.

    Returns a new dict (does not mutate the reconciler's output) with `in_drive`,
    `tally_entry_status`, `client_action` set. If `drive_checker` is None, the
    columns are left blank — the template still renders, ready for that step
    to be wired in later (targeted Drive-verify, Phase 3 task).
    """
    out = dict(row)
    if drive_checker is None:
        return out
    try:
        found = drive_checker(row)
    except Exception as e:
        logger.warning(f"drive_checker failed for row {row.get('supplier_inv') or row.get('invoice_no')}: {e}")
        return out
    if found:
        out["in_drive"] = "Yes"
        out["tally_entry_status"] = "Entry missing in Tally"
    else:
        out["in_drive"] = "No"
        out["client_action"] = "Request invoice from client (pending accountant review)"
    return out


def _annotate_client_action(row: dict) -> dict:
    status = row.get("recon_status")
    action = _CLIENT_ACTION.get(status)
    if action and not row.get("client_action"):
        out = dict(row)
        out["client_action"] = action
        return out
    return row


def write_gstr2b_reconciliation_excel(
    recon_result: dict,
    path: str,
    drive_checker: Optional[Callable[[dict], bool]] = None,
) -> None:
    """
    Writes the reconciliation workbook.

    Args:
        recon_result:  the dict returned by gstr2b_reconciler.reconcile()
                       ({"rows": [...], "extra": [...], "summary": {...}})
        path:          output .xlsx path
        drive_checker: optional callable(row) -> bool, used only on
                       `not_in_books` rows (Section 13.1 gap investigation).
                       Left None until the targeted Drive-verify service exists.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed")

    all_rows = list(recon_result.get("rows", [])) + [
        _annotate_gap_row(r, drive_checker) for r in recon_result.get("extra", [])
    ]
    all_rows = [_annotate_client_action(r) for r in all_rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    cols = labels(GSTR2B_RECONCILIATION_VIEW)
    for col_idx, label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for row_idx, row in enumerate(all_rows, 2):
        values = row_for(GSTR2B_RECONCILIATION_VIEW, row)
        fill = _STATUS_FILL.get(row.get("recon_status"))
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            if fill:
                cell.fill = fill

    for col_idx in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    _write_summary_sheet(wb, recon_result.get("summary", {}))

    wb.save(path)
    logger.info(f"Wrote GSTR-2B reconciliation workbook to {path} ({len(all_rows)} rows)")


def _write_summary_sheet(wb, summary: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    def _header(text, row):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)
        return row + 1

    r = 1
    r = _header("Counts by status", r)
    for status, count in (summary.get("counts") or {}).items():
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2, value=count)
        r += 1

    r += 1
    r = _header("Amounts by status (₹)", r)
    for status, amt in (summary.get("amounts") or {}).items():
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2, value=amt)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="ITC at risk (₹)").font = Font(bold=True)
    ws.cell(row=r, column=2, value=summary.get("itc_at_risk", 0.0))
    r += 1
    ws.cell(row=r, column=1, value="Matched ITC (₹)")
    ws.cell(row=r, column=2, value=summary.get("matched_itc", 0.0))
    r += 1
    ws.cell(row=r, column=1, value="Fuzzy-matched count")
    ws.cell(row=r, column=2, value=summary.get("fuzzy_matched_count", 0))

    r += 2
    r = _header("Rule 36(4) — provisional ITC cap", r)
    rule = summary.get("rule_36_4") or {}
    for label, key in (("Cap (105% of matched)", "cap"), ("Total claimed", "total_claimed"),
                        ("Excess", "excess"), ("Breached?", "breached")):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=rule.get(key))
        r += 1
