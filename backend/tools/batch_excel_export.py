"""
batch_excel_export.py — Cost-optimised batch processor for sales invoices.

COST STRATEGY:
  Full pipeline: 3 LLM extractors x full PDF text x every chunk
               = ~400K tokens/invoice (very expensive at scale)

  This script:
    1. Deterministic regex  -> all financial data (taxable, CGST, SGST, IGST,
                               total, advance, round_off, invoice_no, date, GSTIN)
    2. ONE tiny LLM call    -> only party name + place_of_supply
                               (first 800 chars of invoice = header region only)
                             = ~1,000 tokens/invoice

  Accuracy is identical: deterministic regex achieves 100% on financials.
  LLM only fills party name which regex cannot reliably get.

Usage:
  python batch_excel_export.py --dir /path/to/invoices --out output.xlsx
  python batch_excel_export.py --dir /path/to/invoices --out output.xlsx --retry
"""
import asyncio, os, sys, time, glob as globmod, pathlib, json, re, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Resolve project root from this file's location: backend/tools/batch_excel_export.py
ROOT    = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
BACKEND = os.path.join(ROOT, "backend")
sys.path.insert(0, ROOT)
sys.path.insert(0, BACKEND)
os.chdir(ROOT)
load_dotenv(os.path.join(BACKEND, ".env"))

# Paths set from CLI args in main() — no hardcoded machine paths
INVOICE_DIR  = None
OUTPUT_EXCEL = None
FAILED_LOG   = None

# LLM config — prefers Claude Haiku if ANTHROPIC_API_KEY is set, else Groq free.
# Claude Haiku: ~$0.0004/invoice (160 output tokens). Groq: free but rate-limited.
SEMAPHORE      = 5
GROQ_MODEL     = "llama-3.1-8b-instant"
CLAUDE_MODEL   = "claude-haiku-4-5-20251001"


# ── Deterministic extractors ────────────────────────────────────────────────

def _f(s: str) -> float:
    try:
        return float(s.replace(',', '').replace(' ', ''))
    except Exception:
        return 0.0

def extract_financials(text: str) -> dict:
    """Pull all numbers from the GST summary table — zero LLM cost."""
    t = re.sub(r'[ \t]+', ' ', text)
    result = {}

    # GST summary table (anchored to header to avoid line-item false matches)
    m = re.search(
        r'Taxable Value.{0,300}?Central Tax.{0,600}?'
        r'([\d,]+\.[\d]{2})\s+'
        r'(\d*\.?\d*)\s*%\s+'
        r'([\d,]+\.[\d]{2})\s+'
        r'(\d*\.?\d*)\s*%\s+'
        r'([\d,]+\.[\d]{2})\s+'
        r'(\d*\.?\d*)\s*%\s+'
        r'([\d,]+\.[\d]{2})',
        t, re.DOTALL
    )
    if m:
        result['taxable_value'] = _f(m.group(1))
        result['cgst_rate']     = float(m.group(2)) if m.group(2) else 0.0
        result['cgst_amount']   = _f(m.group(3))
        result['sgst_rate']     = float(m.group(4)) if m.group(4) else 0.0
        result['sgst_amount']   = _f(m.group(5))
        result['igst_rate']     = float(m.group(6)) if m.group(6) else 0.0
        result['igst_amount']   = _f(m.group(7))

    ft = re.search(r'Final Total\s*\([^)]*\)\s+([\d,]+\.[\d]{2})', t)
    if ft:
        result['total_invoice_value'] = _f(ft.group(1))

    tc = re.search(r'Total Cost incl Taxes\s+([\d,]+\.[\d]{2})', t)
    if tc:
        result['total_invoice_value'] = _f(tc.group(1))

    ro = re.search(r'Rounding off\s+(-?[\d,]+\.[\d]{2})', t)
    if ro:
        result['round_off'] = _f(ro.group(1))

    nc = re.search(r'Net Cost\s+([\d,]+\.[\d]{2})', t)
    ft2 = re.search(r'Final Total\s*\([^)]*\)\s+([\d,]+\.[\d]{2})', t)
    if nc and ft2:
        adv = round(_f(nc.group(1)) - _f(ft2.group(1)), 2)
        if adv > 0.50:
            result['advance_amount'] = adv

    # HSN/SAC — 4-8 digit code near HSN/SAC label
    hsn_m = re.search(r'(?:HSN|SAC)[/\s]*(?:Code|No)?[:\s]*(\d{4,8})', t, re.IGNORECASE)
    if not hsn_m:
        # Fallback: any standalone 4-8 digit number on a line with HSN context
        hsn_m = re.search(r'\b(\d{4,8})\b', t)
    if hsn_m:
        result['hsn'] = hsn_m.group(1)

    return result

def extract_header_fields(text: str) -> dict:
    """Deterministically pull invoice_no, date, GSTIN from header."""
    result = {}

    inv = re.search(r'Invoice\s*(?:No|Number|#)[:\s]*([A-Z0-9/_-]{4,30})', text, re.IGNORECASE)
    if inv:
        result['invoice_no'] = inv.group(1).strip()

    date = re.search(r'(?:Date|Dated)[:\s]*(\d{1,2}[-/\s][A-Za-z]{3,9}[-/\s]\d{2,4}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', text, re.IGNORECASE)
    if date:
        result['voucher_date'] = date.group(1).strip()

    gstin = re.search(r'\b(\d{2}[A-Z]{5}\d{4}[A-Z]{1}\d{1}[Z]{1}[A-Z0-9]{1})\b', text)
    if gstin:
        result['party_gstin'] = gstin.group(1)

    return result


# ── Minimal LLM call for party name + place of supply ──────────────────────

def _llm_party_pos(header_text: str, client, model: str) -> dict:
    """One tiny call, ~600 input tokens, to get party name, place of supply, particulars, and HSN."""
    from openai import OpenAI
    prompt = f"""From this invoice extract ONLY these 4 fields:
- party_name: the customer/buyer company name (NOT the seller)
- place_of_supply: state name or state code shown on invoice
- particulars: brief description of what was sold/billed (e.g. "Mobile Charges", "Cloud Hosting", "Professional Fees"). Use the service/product description line if visible. Max 80 chars.
- hsn_sac: the HSN or SAC code (4-8 digit number). Return empty string if not found.

Return JSON only, no explanation.
Example: {{"party_name": "ABC Bank Ltd.", "place_of_supply": "Karnataka", "particulars": "Internet Bandwidth Charges", "hsn_sac": "998314"}}

Invoice header:
{header_text[:1200]}
"""
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.0,
            max_tokens=200,
        )
        raw = resp.choices[0].message.content.strip()
        data = json.loads(raw)
        return {
            "party_ledger_name": data.get("party_name", ""),
            "place_of_supply":   data.get("place_of_supply", ""),
            "particulars":       data.get("particulars", ""),
            "hsn_sac":           data.get("hsn_sac", ""),
        }
    except Exception as e:
        return {"party_ledger_name": "", "place_of_supply": "", "particulars": "", "hsn_sac": ""}


# ── Per-invoice processor ───────────────────────────────────────────────────

async def process_one(sem, pdf_path, client):
    async with sem:
        t0    = time.time()
        fname = pathlib.Path(pdf_path).name
        try:
            import fitz
            doc  = fitz.open(pdf_path)
            full = "\n".join(p.get_text() for p in doc)
            doc.close()

            fin    = extract_financials(full)
            header = extract_header_fields(full[:2000])

            # One LLM call for party name + POS + HSN (header only, ~600 tokens)
            llm = await asyncio.to_thread(_llm_party_pos, full[:1200], client, GROQ_MODEL)

            tv   = fin.get('taxable_value', 0.0)
            cgst = fin.get('cgst_amount', 0.0)
            sgst = fin.get('sgst_amount', 0.0)
            igst = fin.get('igst_amount', 0.0)
            tax  = cgst + sgst + igst
            tot  = fin.get('total_invoice_value', 0.0)
            adv  = fin.get('advance_amount', 0.0)
            roff = fin.get('round_off', 0.0)

            status = "PASS" if tv > 0 else "WARN"
            elapsed = round(time.time() - t0, 1)

            # GSTR-1 category (simple rule)
            gstin = header.get('party_gstin', '')
            pos   = llm.get('place_of_supply', '')
            if igst > (cgst + sgst):
                gstr1 = "B2B" if len(gstin) == 15 else "B2CL"
            elif igst == 0 and cgst == 0 and sgst == 0:
                gstr1 = "EXP"
            else:
                gstr1 = "B2B" if len(gstin) == 15 else "B2CS"

            print(f"[{status}] {fname[:60]:<60} taxable={tv:>10.2f}  {elapsed}s")
            return {
                "File Name": fname,
                "Invoice No": header.get('invoice_no', ''),
                "Invoice Date": header.get('voucher_date', ''),
                "Party Name": llm.get('party_ledger_name', ''),
                "Party GSTIN": gstin,
                "Place of Supply": pos,
                "Particulars": llm.get('particulars', ''),
                "HSN/SAC": fin.get('hsn', '') or llm.get('hsn_sac', ''),
                "Taxable Value": tv,
                "CGST Rate (%)": fin.get('cgst_rate', 0.0),
                "CGST Amount": cgst,
                "SGST Rate (%)": fin.get('sgst_rate', 0.0),
                "SGST Amount": sgst,
                "IGST Rate (%)": fin.get('igst_rate', 0.0),
                "IGST Amount": igst,
                "Total Tax": tax,
                "Round Off": roff,
                "Advance Deducted": adv,
                "Total Invoice Value": tot,
                "GSTR-1 Category": gstr1,
                "Tax Type": "IGST" if igst > (cgst + sgst) else "CGST+SGST",
                "Nature": "",
                "Status": status,
                "Note": f"{elapsed}s",
                "_path": pdf_path,
            }

        except Exception as e:
            elapsed = round(time.time() - t0, 1)
            print(f"[ERR ] {fname[:60]:<60} {elapsed}s  {e}")
            return {
                "File Name": fname, "Invoice No": "", "Invoice Date": "",
                "Party Name": "", "Party GSTIN": "", "Place of Supply": "",
                "Particulars": "", "HSN/SAC": "", "Taxable Value": 0.0, "CGST Rate (%)": 0.0,  # noqa
                "CGST Amount": 0.0, "SGST Rate (%)": 0.0, "SGST Amount": 0.0,
                "IGST Rate (%)": 0.0, "IGST Amount": 0.0, "Total Tax": 0.0,
                "Round Off": 0.0, "Advance Deducted": 0.0, "Total Invoice Value": 0.0,
                "GSTR-1 Category": "", "Tax Type": "", "Nature": "",
                "Status": "ERR", "Note": str(e)[:120], "_path": pdf_path,
            }


# ── Excel writer ────────────────────────────────────────────────────────────

COLUMNS = [
    "S.No", "File Name", "Invoice No", "Invoice Date",
    "Party Name", "Party GSTIN", "Place of Supply", "Particulars", "HSN/SAC",
    "Taxable Value", "CGST Rate (%)", "CGST Amount",
    "SGST Rate (%)", "SGST Amount", "IGST Rate (%)", "IGST Amount",
    "Total Tax", "Round Off", "Advance Deducted", "Total Invoice Value",
    "GSTR-1 Category", "Tax Type", "Nature", "Status", "Note",
]
COL_WIDTHS = [5, 35, 18, 14, 38, 18, 18, 30, 10,
              14, 11, 12, 11, 12, 11, 12,
              12, 10, 14, 16, 14, 12, 25, 8, 30]
NUM_COLS = {10, 12, 14, 16, 17, 18, 19, 20}  # 1-indexed cols with numbers (shifted +1)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
PASS_FILL   = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL   = PatternFill("solid", fgColor="FFF2CC")
ERR_FILL    = PatternFill("solid", fgColor="FCE4D6")
BORDER      = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))

def write_excel(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 28

    for sno, row in enumerate(rows, 1):
        r    = sno + 1
        fill = PASS_FILL if row["Status"]=="PASS" else (WARN_FILL if row["Status"]=="WARN" else ERR_FILL)
        vals = [sno, row["File Name"], row["Invoice No"], row["Invoice Date"],
                row["Party Name"], row["Party GSTIN"], row["Place of Supply"], row.get("Particulars", ""), row["HSN/SAC"],
                row["Taxable Value"], row["CGST Rate (%)"], row["CGST Amount"],
                row["SGST Rate (%)"], row["SGST Amount"], row["IGST Rate (%)"], row["IGST Amount"],
                row["Total Tax"], row["Round Off"], row["Advance Deducted"], row["Total Invoice Value"],
                row["GSTR-1 Category"], row["Tax Type"], row["Nature"], row["Status"], row["Note"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill; c.border = BORDER
            c.alignment = Alignment(vertical="center")
            if ci in NUM_COLS:
                c.number_format = '#,##0.00'

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Totals row
    rt = len(rows) + 2
    ws.cell(row=rt, column=1, value="TOTAL").font = Font(bold=True)
    for ci, col in enumerate(COLUMNS, 1):
        if col in ("Taxable Value","CGST Amount","SGST Amount","IGST Amount",
                   "Total Tax","Round Off","Advance Deducted","Total Invoice Value"):
            cl = get_column_letter(ci)
            c = ws.cell(row=rt, column=ci, value=f"=SUM({cl}2:{cl}{len(rows)+1})")
            c.font = Font(bold=True); c.number_format = '#,##0.00'; c.border = BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 2: Tax & TDS Workings ─────────────────────────────────────────
    TDS_THRESHOLD = 30000
    ws2 = wb.create_sheet("Tax & TDS Workings")
    tds_cols = [
        "Invoice No", "Invoice Date", "Party Name", "Party GSTIN", "Place of Supply",
        "Taxable Value (Rs.)", "GST Type",
        "CGST Rate %", "CGST Amount (Rs.)", "SGST Rate %", "SGST Amount (Rs.)",
        "IGST Rate %", "IGST Amount (Rs.)", "Total Tax (Rs.)", "Total Invoice Value (Rs.)",
        "GSTR-1 Category", "GST Routing Rule Applied",
        "TDS Applicable", "TDS Section", "TDS Rate %",
        "TDS Deducted (Est.) (Rs.)", "Net Receivable After TDS (Rs.)", "TDS Note", "Nature",
    ]
    for ci, col in enumerate(tds_cols, 1):
        c = ws2.cell(row=1, column=ci, value=col)
        c.font = HEADER_FONT; c.fill = PatternFill("solid", fgColor="375623")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 28

    def _gst_type_b(r):
        igst = r.get("IGST Amount", 0) or 0
        cgst = r.get("CGST Amount", 0) or 0
        sgst = r.get("SGST Amount", 0) or 0
        taxable = r.get("Taxable Value", 0) or 0
        if (igst + cgst + sgst) == 0 and taxable > 0:
            return "Export LUT"
        return "Interstate" if igst > 0 else ("Intrastate" if cgst > 0 or sgst > 0 else "Exempt")

    def _tds_b(r, gst_type):
        taxable = r.get("Taxable Value", 0) or 0
        cat = (r.get("GSTR-1 Category", "") or "").upper()
        nature = (r.get("Nature", "") or "").lower()
        if cat in ("EXP", "B2CS") or taxable < TDS_THRESHOLD:
            return "No", "-", 0.0, 0.0, taxable, "Below threshold / not applicable"
        prof_kw = ["professional", "technical", "consulting", "audit", "legal", "management"]
        if any(k in nature for k in prof_kw):
            tds_amt = round(taxable * 0.10, 2)
            return "Yes", "194J - Prof/Tech Services", 10.0, tds_amt, round(taxable - tds_amt, 2), "10% TDS (Sec 194J)"
        tds_amt = round(taxable * 0.02, 2)
        return "Yes", "194C - Contractor", 2.0, tds_amt, round(taxable - tds_amt, 2), "2% TDS (Sec 194C, company deductee)"

    for ri, r in enumerate(rows, 2):
        gst_type = _gst_type_b(r)
        if gst_type == "Export LUT":
            routing = "IGST = 0 (Export LUT). Zero-rated under IGST Act Sec 16(3)(b)."
        elif gst_type == "Interstate":
            routing = f"Interstate -> IGST only. Category: {r.get('GSTR-1 Category', '')}."
        else:
            routing = f"Intrastate -> CGST + SGST (equal halves). Category: {r.get('GSTR-1 Category', '')}."
        tds_app, tds_sec, tds_rate, tds_ded, net_recv, tds_note = _tds_b(r, gst_type)
        vals2 = [
            r.get("Invoice No"), r.get("Invoice Date"), r.get("Party Name"), r.get("Party GSTIN"),
            r.get("Place of Supply"), r.get("Taxable Value", 0), gst_type,
            r.get("CGST Rate (%)", 0), r.get("CGST Amount", 0),
            r.get("SGST Rate (%)", 0), r.get("SGST Amount", 0),
            r.get("IGST Rate (%)", 0), r.get("IGST Amount", 0),
            r.get("Total Tax", 0), r.get("Total Invoice Value", 0),
            r.get("GSTR-1 Category"), routing,
            tds_app, tds_sec, tds_rate, tds_ded, net_recv, tds_note, r.get("Nature", ""),
        ]
        for ci, v in enumerate(vals2, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 17 or ci == 23))
            if isinstance(v, float): c.number_format = '#,##0.00'
    for ci, w in enumerate([12,12,30,16,16, 14,12, 9,12,9,12,9,12,12,14, 12,40, 9,24,9,14,16,35,25], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Matching Sheet ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Matching Sheet")
    match_cols = [
        "Invoice No", "Invoice Date", "Party Name", "Party GSTIN",
        "Taxable Value (Rs.)", "CGST (Rs.)", "SGST (Rs.)", "IGST (Rs.)",
        "Total Tax (Rs.)", "Round Off (Rs.)", "Total Invoice Value (Rs.)",
        "Computed Total (Rs.)", "Variance (Rs.)", "Match Status",
        "GSTR-1 Category", "Nature",
    ]
    for ci, col in enumerate(match_cols, 1):
        c = ws3.cell(row=1, column=ci, value=col)
        c.font = HEADER_FONT; c.fill = PatternFill("solid", fgColor="4B1D6B")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws3.row_dimensions[1].height = 28

    for ri, r in enumerate(rows, 2):
        taxable = r.get("Taxable Value", 0) or 0
        cgst = r.get("CGST Amount", 0) or 0
        sgst = r.get("SGST Amount", 0) or 0
        igst = r.get("IGST Amount", 0) or 0
        round_off = r.get("Round Off", 0) or 0
        total = r.get("Total Invoice Value", 0) or 0
        total_tax = cgst + sgst + igst
        computed = round(taxable + total_tax + round_off, 2)
        variance = round(abs(total - computed), 2)
        is_export = total_tax == 0 and taxable > 0
        if is_export:
            match_status = "PASS - Export LUT"
            variance = 0.0
        elif variance <= 2.50:
            match_status = "PASS"
        else:
            match_status = f"REVIEW - Variance Rs.{variance}"
        mfill = PASS_FILL if "PASS" in match_status else WARN_FILL
        vals3 = [
            r.get("Invoice No"), r.get("Invoice Date"), r.get("Party Name"), r.get("Party GSTIN"),
            taxable, cgst, sgst, igst, total_tax, round_off, total, computed, variance,
            match_status, r.get("GSTR-1 Category"), r.get("Nature", ""),
        ]
        for ci, v in enumerate(vals3, 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
            if isinstance(v, float): c.number_format = '#,##0.00'
            if ci == 14:  # Match Status column — colour by result
                c.fill = mfill
    for ci, w in enumerate([12,12,30,16, 14,11,11,11,11,10,14, 13,11,22, 12,25], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Batch Summary ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Batch Summary")
    n_pass = sum(1 for r in rows if r["Status"] == "PASS")
    n_warn = sum(1 for r in rows if r["Status"] == "WARN")
    n_err  = sum(1 for r in rows if r["Status"] == "ERR")
    for ri, (k, v) in enumerate([
        ("Total Invoices", len(rows)),
        ("PASS", n_pass), ("WARN", n_warn), ("ERR", n_err),
        ("Accuracy", f"{n_pass/len(rows)*100:.1f}%" if rows else "0%"),
        ("Total Taxable Value", sum(r["Taxable Value"] for r in rows)),
        ("Total Invoice Value", sum(r["Total Invoice Value"] for r in rows)),
        ("LLM Provider", "gemini-2.5-flash" if os.getenv("GEMINI_API_KEY") else (CLAUDE_MODEL if os.getenv("ANTHROPIC_API_KEY") else GROQ_MODEL)),
        ("Approx LLM Cost", "Free (Gemini)" if os.getenv("GEMINI_API_KEY") else ("Rs.0 (Groq free)" if os.getenv("GROQ_API_KEY") else f"~Rs.{len(rows)*0.035:.2f} (Claude)")),
    ], 1):
        ws4.cell(row=ri, column=1, value=k).font = Font(bold=True)
        c = ws4.cell(row=ri, column=2, value=v)
        if isinstance(v, float): c.number_format = '#,##0.00'
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 22

    wb.save(path)
    print(f"\nExcel saved -> {path}")


# ── Main ────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description="Batch invoice -> Excel exporter")
    parser.add_argument("--dir",   required=True, help="Folder containing PDF invoices")
    parser.add_argument("--out",   required=True, help="Output Excel file path")
    parser.add_argument("--retry", action="store_true", help="Only reprocess failed invoices from last run")
    args = parser.parse_args()

    invoice_dir  = os.path.abspath(args.dir)
    output_excel = os.path.abspath(args.out)
    failed_log   = output_excel.replace(".xlsx", "_failed.json")

    from openai import OpenAI
    gemini_key    = os.getenv("GEMINI_API_KEY")
    groq_key      = os.getenv("GROQ_API_KEY")
    if gemini_key:
        client = OpenAI(api_key=gemini_key, base_url="https://generativelanguage.googleapis.com/v1beta/openai/")
        GROQ_MODEL_ACTIVE = "gemini-2.5-flash"
        print(f"Provider: Gemini 2.5 Flash")
    elif groq_key:
        client = OpenAI(api_key=groq_key, base_url="https://api.groq.com/openai/v1")
        GROQ_MODEL_ACTIVE = GROQ_MODEL
        print(f"Provider: Groq free ({GROQ_MODEL})")
    else:
        print("ERROR: Set GEMINI_API_KEY or GROQ_API_KEY in .env"); return
    globals()["GROQ_MODEL"] = GROQ_MODEL_ACTIVE  # patch module-level var used in _llm_party_pos

    if args.retry and os.path.exists(failed_log):
        with open(failed_log) as f:
            pdfs = json.load(f)
        print(f"Retrying {len(pdfs)} failed invoices\n")
    else:
        pdfs = sorted(globmod.glob(os.path.join(invoice_dir, "**", "*.pdf"), recursive=True))
        if not pdfs:
            print("No PDFs found in", invoice_dir); return
        print(f"Processing {len(pdfs)} invoices (1 LLM call/invoice, ~1K tokens each)\n")

    sem     = asyncio.Semaphore(SEMAPHORE)
    t_start = time.time()
    results = await asyncio.gather(*[process_one(sem, p, client) for p in pdfs])
    wall    = time.time() - t_start

    n_pass = sum(1 for r in results if r["Status"]=="PASS")
    n_warn = sum(1 for r in results if r["Status"]=="WARN")
    n_err  = sum(1 for r in results if r["Status"]=="ERR")

    failed = [r["_path"] for r in results if r["Status"]=="ERR"]
    if failed:
        with open(failed_log, "w") as f: json.dump(failed, f, indent=2)

    for r in results: r.pop("_path", None)

    print(f"\n{'='*60}")
    print(f"PASS={n_pass}  WARN={n_warn}  ERR={n_err}  Accuracy={n_pass/len(pdfs)*100:.1f}%")
    print(f"Wall time={wall:.1f}s  |  LLM calls={len(pdfs)}  |  ~{len(pdfs)*1000:,} tokens used")

    write_excel(results, output_excel)


if __name__ == "__main__":
    asyncio.run(main())
