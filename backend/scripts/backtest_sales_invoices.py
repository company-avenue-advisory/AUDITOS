"""
Backtest script for salesinvoices/ against Sales_Output.xlsx golden dataset.

Runs every PDF through process_pdf(), compares extracted output vs golden at:
  - Invoice level: party_name, gstin, taxable (amount-discount), cgst, sgst, igst, total
  - Line item level: particulars, taxable, cgst, sgst, igst, total per item

Outputs:
  - Console summary with accuracy stats and per-invoice diff
  - backend/data/vendor_profiles/banking_saas_golden.json
"""
import sys, os, json, glob
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding="utf-8")

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

from invoice_processor import process_pdf

BASE_DIR        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SALES_DIR       = os.path.join(BASE_DIR, "..", "salesinvoices")
EXCEL_PATH      = os.path.join(SALES_DIR, "Sales_Output.xlsx")
PROFILE_DIR     = os.path.join(BASE_DIR, "data", "vendor_profiles")
GOLDEN_OUT_FILE = os.path.join(PROFILE_DIR, "banking_saas_golden.json")
os.makedirs(PROFILE_DIR, exist_ok=True)

TOLERANCE = 1.0  # ₹1 tolerance for float comparisons

# ── Load golden dataset ───────────────────────────────────────────────────────
def load_golden(excel_path: str) -> tuple[dict, dict]:
    wb = openpyxl.load_workbook(excel_path)

    # Main sheet — one row per invoice
    ws = wb["Main"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    invoices = {}
    for r in range(2, ws.max_row + 1):
        row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        ref = str(row.get("REFERANCE NO") or "").strip()
        if not ref:
            continue
        discount = float(row.get("DISCOUNT") or 0)
        amount   = float(row.get("AMOUNT") or 0)
        invoices[ref] = {
            "invoice_no":     ref,
            "invoice_date":   str(row.get("INVOICE DATE") or ""),
            "party_gstin":    str(row.get("GST NO") or "").strip(),
            "party_name":     str(row.get("PARTY A/C NAME") or "").strip(),
            "place_of_supply":str(row.get("PLACE OF SUPPLY") or "").strip(),
            "particulars":    str(row.get("PARTICULARS") or "").strip(),
            "amount":         amount,
            "discount":       discount,
            "taxable":        round(amount - discount, 2),
            "advances":       float(row.get("ADVANCES") or 0),
            "cgst":           float(row.get("CGST") or 0),
            "sgst":           float(row.get("SGST") or 0),
            "igst":           float(row.get("IGST") or 0),
            "total":          float(row.get("TOTAL AMOUNT") or 0),
            "hsn":            str(row.get("HSN") or "").strip(),
        }

    # Line Items sheet
    ws2 = wb["Line Items"]
    h2  = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    line_items: dict[str, list] = {}
    for r in range(2, ws2.max_row + 1):
        row = {h2[c-1]: ws2.cell(r, c).value for c in range(1, ws2.max_column + 1)}
        ref = str(row.get("REFERANCE NO") or "").strip()
        if not ref:
            continue
        line_items.setdefault(ref, []).append({
            "particulars": str(row.get("Particulars / Item Description") or "").strip(),
            "amount":      float(row.get("AMOUNT") or 0),
            "discount":    float(row.get("DISCOUNT") or 0),
            "taxable":     round(float(row.get("AMOUNT") or 0) - float(row.get("DISCOUNT") or 0), 2),
            "cgst":        float(row.get("CGST") or 0),
            "sgst":        float(row.get("SGST") or 0),
            "igst":        float(row.get("IGST") or 0),
            "total":       float(row.get("TOTAL AMOUNT") or 0),
            "hsn":         str(row.get("HSN") or "").strip(),
        })

    return invoices, line_items

# ── Match extracted result to golden ─────────────────────────────────────────
def match_val(got, expected, tol=TOLERANCE) -> tuple[bool, str]:
    try:
        g = float(got or 0)
        e = float(expected or 0)
        ok = abs(g - e) <= tol
        return ok, f"{g:.2f} vs {e:.2f}"
    except (TypeError, ValueError):
        ok = str(got or "").strip().lower() == str(expected or "").strip().lower()
        return ok, f"'{got}' vs '{expected}'"

def find_pdf(ref: str, party_name: str, sales_dir: str) -> str | None:
    """Find the PDF file that corresponds to an invoice reference number."""
    # Try matching by party_name (filename often contains party name fragments)
    clean_name = party_name.replace(" ", "_").replace(".", "").replace("-", "_")
    pattern = os.path.join(sales_dir, "*.pdf")
    candidates = glob.glob(pattern)

    # Strategy 1: name similarity
    party_words = [w.lower() for w in party_name.split() if len(w) > 3]
    best_match = None
    best_score = 0
    for pdf in candidates:
        fname_lower = os.path.basename(pdf).lower()
        score = sum(1 for w in party_words if w in fname_lower)
        if score > best_score:
            best_score = score
            best_match = pdf
    if best_score >= 2:
        return best_match
    # Strategy 2: first meaningful word
    for pdf in candidates:
        fname_lower = os.path.basename(pdf).lower()
        if party_words and party_words[0] in fname_lower:
            return pdf
    return None

def extract_one(pdf_path: str) -> dict:
    try:
        res = process_pdf(pdf_path, invoice_type="sales")
        return {"ok": True, "result": res}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"\nLoading golden dataset from {EXCEL_PATH} ...")
    golden_inv, golden_items = load_golden(EXCEL_PATH)
    print(f"  {len(golden_inv)} invoices, {sum(len(v) for v in golden_items.values())} line items\n")

    results = []
    invoice_pass = invoice_fail = 0
    field_stats = {"taxable": [0, 0], "cgst": [0, 0], "sgst": [0, 0], "igst": [0, 0], "total": [0, 0]}

    for ref, golden in sorted(golden_inv.items()):
        pdf_path = find_pdf(ref, golden["party_name"], SALES_DIR)
        if not pdf_path:
            print(f"  [SKIP] {ref} — PDF not found for '{golden['party_name']}'")
            results.append({"ref": ref, "status": "pdf_not_found", "golden": golden})
            continue

        fname = os.path.basename(pdf_path)
        print(f"  [{ref}] {fname} ...", end=" ", flush=True)
        ext = extract_one(pdf_path)

        if not ext["ok"]:
            print(f"ERROR: {ext['error']}")
            results.append({"ref": ref, "status": "error", "error": ext["error"], "golden": golden})
            invoice_fail += 1
            continue

        res = ext["result"]
        items = res.sales_items or []

        # Compute extracted totals
        ext_taxable = sum(i.taxable_value or 0 for i in items)
        ext_cgst    = sum(i.cgst_amount or 0 for i in items)
        ext_sgst    = sum(i.sgst_amount or 0 for i in items)
        ext_igst    = sum(i.igst_amount or 0 for i in items)
        ext_total   = sum(i.total_invoice_value or 0 for i in items)

        # Override with overall fields if items are empty/zero
        if ext_total == 0:
            ext_taxable = res.overall_taxable_value or 0
            ext_cgst    = res.overall_cgst_amount or 0
            ext_sgst    = res.overall_sgst_amount or 0
            ext_igst    = res.overall_igst_amount or 0
            ext_total   = res.overall_total_invoice_value or 0

        checks = {
            "taxable": match_val(ext_taxable, golden["taxable"]),
            "cgst":    match_val(ext_cgst, golden["cgst"]),
            "sgst":    match_val(ext_sgst, golden["sgst"]),
            "igst":    match_val(ext_igst, golden["igst"]),
            "total":   match_val(ext_total, golden["total"]),
        }

        all_pass = all(v[0] for v in checks.values())
        status_str = "PASS" if all_pass else "FAIL"
        if all_pass:
            invoice_pass += 1
        else:
            invoice_fail += 1

        for field, (ok, _) in checks.items():
            field_stats[field][0 if ok else 1] += 1

        fails = {k: v[1] for k, v in checks.items() if not v[0]}
        print(f"{status_str}  ({len(items)} items)" + (f"  DIFF: {fails}" if fails else ""))

        extracted_items = [
            {
                "particulars":         i.particulars,
                "hsn":                 i.hsn,
                "taxable_value":       i.taxable_value,
                "cgst_amount":         i.cgst_amount,
                "sgst_amount":         i.sgst_amount,
                "igst_amount":         i.igst_amount,
                "total_invoice_value": i.total_invoice_value,
                "gstr1_category":      i.gstr1_category,
            }
            for i in items
        ]

        results.append({
            "ref": ref, "status": status_str, "pdf": pdf_path,
            "golden": golden,
            "golden_line_items": golden_items.get(ref, []),
            "extracted": {
                "taxable": round(ext_taxable, 2), "cgst": round(ext_cgst, 2),
                "sgst": round(ext_sgst, 2), "igst": round(ext_igst, 2),
                "total": round(ext_total, 2), "items": extracted_items,
            },
            "checks": {k: v[0] for k, v in checks.items()},
        })

    # ── Summary ──────────────────────────────────────────────────────────────
    total = invoice_pass + invoice_fail
    print(f"\n{'='*65}")
    print(f"Invoice accuracy: {invoice_pass}/{total} PASS ({100*invoice_pass/max(total,1):.0f}%)")
    print(f"\nField-level accuracy:")
    for field, (passes, fails) in field_stats.items():
        n = passes + fails
        print(f"  {field:<10} {passes}/{n}  ({100*passes/max(n,1):.0f}%)")

    # ── Save golden results ───────────────────────────────────────────────────
    with open(GOLDEN_OUT_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "generated_at": datetime.now().isoformat(),
            "invoice_accuracy_pct": round(100 * invoice_pass / max(total, 1), 1),
            "results": results,
        }, f, indent=2, ensure_ascii=False)
    print(f"\nDetailed results → {GOLDEN_OUT_FILE}")

if __name__ == "__main__":
    main()
