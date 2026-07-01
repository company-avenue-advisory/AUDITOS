"""
Training + validation script for One Stack Solution invoices.
Runs 15 selected bank invoices through process_pdf(), compares vs golden,
and writes backend/data/vendor_profiles/06AAGCR4375J2Z1.json.
"""
import sys, os, json, glob
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding="utf-8")

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

import openpyxl
from invoice_processor import process_pdf

BASE_DIR     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SALES_DIR    = os.path.join(BASE_DIR, "..", "salesinvoices")
EXCEL_PATH   = os.path.join(SALES_DIR, "Sales_Output.xlsx")
PROFILE_DIR  = os.path.join(BASE_DIR, "data", "vendor_profiles")
PROFILE_FILE = os.path.join(PROFILE_DIR, "06AAGCR4375J2Z1.json")
GOLDEN_FILE  = os.path.join(PROFILE_DIR, "06AAGCR4375J2Z1_golden.json")
os.makedirs(PROFILE_DIR, exist_ok=True)

TOLERANCE = 2.0

# Selected 15 diverse invoices
TRAINING_REFS = [
    'HR26051012', 'MH26051007', 'MH26051046', 'MH26051067', 'MH26051059',
    'MH26051062', 'MH26051022', 'MH26051074', 'MH26051072', 'MH26051041',
    'HR26051018', 'MH26051001', 'MH26051076', 'MH26051090', 'MH26051099',
]

# ── Load golden ───────────────────────────────────────────────────────────────
def load_golden():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Main']
    h  = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    inv = {}
    for r in range(2, ws.max_row+1):
        row = {h[c-1]: ws.cell(r,c).value for c in range(1, ws.max_column+1)}
        ref = str(row.get('REFERANCE NO') or '').strip()
        if ref not in TRAINING_REFS:
            continue
        disc = float(row.get('DISCOUNT') or 0)
        amt  = float(row.get('AMOUNT') or 0)
        inv[ref] = {
            'ref':     ref,
            'party':   str(row.get('PARTY A/C NAME') or '').strip(),
            'gstin':   str(row.get('GST NO') or '').strip(),
            'taxable': round(amt - disc, 2),
            'amount':  amt,
            'discount':disc,
            'advance': float(row.get('ADVANCES') or 0),
            'cgst':    float(row.get('CGST') or 0),
            'sgst':    float(row.get('SGST') or 0),
            'igst':    float(row.get('IGST') or 0),
            'total':   float(row.get('TOTAL AMOUNT') or 0),
        }

    ws2 = wb['Line Items']
    h2  = [ws2.cell(1,c).value for c in range(1, ws2.max_column+1)]
    items = {}
    for r in range(2, ws2.max_row+1):
        row = {h2[c-1]: ws2.cell(r,c).value for c in range(1, ws2.max_column+1)}
        ref = str(row.get('REFERANCE NO') or '').strip()
        if ref not in TRAINING_REFS:
            continue
        items.setdefault(ref, []).append({
            'particulars': str(row.get('Particulars / Item Description') or '').strip(),
            'amount':      float(row.get('AMOUNT') or 0),
            'discount':    float(row.get('DISCOUNT') or 0),
            'taxable':     round(float(row.get('AMOUNT') or 0) - float(row.get('DISCOUNT') or 0), 2),
            'cgst':        float(row.get('CGST') or 0),
            'sgst':        float(row.get('SGST') or 0),
            'igst':        float(row.get('IGST') or 0),
            'total':       float(row.get('TOTAL AMOUNT') or 0),
            'hsn':         str(row.get('HSN') or '').strip(),
        })
    return inv, items

# ── Find PDF ──────────────────────────────────────────────────────────────────
def find_pdf(ref: str, party: str) -> str:
    pdfs = glob.glob(os.path.join(SALES_DIR, "*.pdf"))
    words = [w.lower() for w in party.split() if len(w) > 3]
    best, best_score = None, 0
    for pdf in pdfs:
        fn = os.path.basename(pdf).lower()
        score = sum(1 for w in words if w in fn)
        if score > best_score:
            best_score, best = score, pdf
    return best if best_score >= 2 else None

# ── Compare ───────────────────────────────────────────────────────────────────
def ok(got, exp, tol=TOLERANCE):
    try:
        return abs(float(got or 0) - float(exp or 0)) <= tol
    except:
        return str(got).strip().lower() == str(exp).strip().lower()

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"\nLoading golden for {len(TRAINING_REFS)} invoices...")
    golden, golden_items = load_golden()
    print(f"  Loaded {len(golden)} invoice summaries\n{'='*65}")

    results = []
    passed = failed = 0

    for ref in TRAINING_REFS:
        g = golden.get(ref)
        if not g:
            print(f"  [{ref}] not in golden — skip")
            continue

        pdf = find_pdf(ref, g['party'])
        if not pdf:
            print(f"  [{ref}] PDF not found for '{g['party']}'")
            results.append({'ref': ref, 'status': 'pdf_not_found', 'golden': g})
            continue

        print(f"  [{ref}] {os.path.basename(pdf)[:45]} ...", end=' ', flush=True)
        try:
            res = process_pdf(pdf, invoice_type='sales')
        except Exception as e:
            print(f"ERROR: {e}")
            results.append({'ref': ref, 'status': 'error', 'error': str(e), 'golden': g})
            failed += 1
            continue

        items = res.sales_items or []
        # Golden was produced by the old script which summed per-row billing table values.
        # Per-item sums are the correct source; gst_table (invoice summary) reflects
        # post-advance/aggregated totals that differ from the per-row billing entries.
        # Total = gross billing amount + taxes = sum(taxable + discount + cgst + sgst + igst)
        # because the golden's TOTAL AMOUNT column = Amount(gross) + taxes, discount separate.
        ext_taxable = sum(i.taxable_value or 0 for i in items) or res.overall_taxable_value or 0
        ext_cgst    = sum(i.cgst_amount   or 0 for i in items) or res.overall_cgst_amount   or 0
        ext_sgst    = sum(i.sgst_amount   or 0 for i in items) or res.overall_sgst_amount   or 0
        ext_igst    = sum(i.igst_amount   or 0 for i in items) or res.overall_igst_amount   or 0
        ext_total   = sum(
            (i.taxable_value or 0) + (i.discount or 0) +
            (i.cgst_amount or 0) + (i.sgst_amount or 0) + (i.igst_amount or 0)
            for i in items
        ) or res.overall_total_invoice_value or 0

        checks = {
            'taxable': ok(ext_taxable, g['taxable']),
            'cgst':    ok(ext_cgst,    g['cgst']),
            'sgst':    ok(ext_sgst,    g['sgst']),
            'igst':    ok(ext_igst,    g['igst']),
            'total':   ok(ext_total,   g['total']),
        }
        all_ok = all(checks.values())
        if all_ok: passed += 1
        else:       failed += 1

        fails = {k: f"{float(v2 or 0):.2f} vs {float(g[k] or 0):.2f}"
                 for k, v2 in [('taxable',ext_taxable),('cgst',ext_cgst),
                                ('sgst',ext_sgst),('igst',ext_igst),('total',ext_total)]
                 if not checks[k]}
        tag = 'PASS' if all_ok else 'FAIL'
        print(f"{tag}  ({len(items)} items)" + (f"  DIFF: {fails}" if fails else ''))

        results.append({
            'ref': ref, 'status': tag, 'pdf': pdf,
            'golden': g,
            'golden_items': golden_items.get(ref, []),
            'extracted': {
                'taxable': round(ext_taxable,2), 'cgst': round(ext_cgst,2),
                'sgst': round(ext_sgst,2),       'igst': round(ext_igst,2),
                'total': round(ext_total,2),
                'items': [{'particulars': i.particulars, 'taxable': i.taxable_value,
                           'discount': i.discount,
                           'cgst': i.cgst_amount, 'sgst': i.sgst_amount,
                           'igst': i.igst_amount, 'total': i.total_invoice_value,
                           'hsn': i.hsn, 'category': i.gstr1_category} for i in items],
            },
            'checks': checks,
        })

    total = passed + failed
    print(f"\n{'='*65}")
    print(f"Result: {passed}/{total} PASS ({100*passed//max(total,1)}%)")

    # Field-level stats
    for field in ['taxable','cgst','sgst','igst','total']:
        p = sum(1 for r in results if r.get('checks',{}).get(field))
        n = sum(1 for r in results if 'checks' in r)
        print(f"  {field:<10} {p}/{n}  ({100*p//max(n,1)}%)")

    # Save golden results
    with open(GOLDEN_FILE, 'w', encoding='utf-8') as f:
        json.dump({'generated_at': datetime.now().isoformat(), 'results': results}, f, indent=2, ensure_ascii=False)

    # Build vendor profile from passed results
    ok_results = [r for r in results if r.get('status') == 'PASS' and r.get('extracted',{}).get('items')]
    hsn_codes = set()
    for r in results:
        for it in r.get('extracted',{}).get('items',[]):
            if it.get('hsn'): hsn_codes.add(str(it['hsn']))

    examples = []
    seen_types = set()
    for r in ok_results:
        items = r['extracted']['items']
        if not items: continue
        g = r['golden']
        tax_type = 'igst' if g['igst'] > 0 else 'cgst_sgst'
        disc_type = 'with_discount' if g['discount'] > 0 else 'no_discount'
        key = f"{tax_type}_{disc_type}"
        if key in seen_types or len(examples) >= 3: continue
        seen_types.add(key)
        it = items[0]
        examples.append({
            'invoice_type': key,
            'ref': r['ref'],
            'output': {
                'particulars': it['particulars'],
                'hsn': it['hsn'],
                'taxable_value': it['taxable'],
                'cgst_amount': it['cgst'],
                'sgst_amount': it['sgst'],
                'igst_amount': it['igst'],
                'total_invoice_value': it['total'],
                'gstr1_category': it['category'],
            }
        })

    profile = {
        'gstin':            '06AAGCR4375J2Z1',
        'name':             'ONE STACK SOLUTION PRIVATE LIMITED',
        'state':            'Haryana',
        'state_code':       '06',
        'is_domestic_only': True,
        'invoice_number_prefix': ['MH260510', 'HR260510'],
        'hsn_codes':        sorted(hsn_codes),
        'tax_rule':         'interstate_igst_18_intrastate_cgst_sgst_18',
        'domestic_tax_rate': 18,
        'product_type':     'SaaS / Banking Software',
        'prompt_hints': [
            'KNOWN VENDOR: ONE STACK SOLUTION PRIVATE LIMITED (GSTIN 06AAGCR4375J2Z1, Haryana) — DOMESTIC B2B INVOICE ONLY. Never set gstr1_category to EXPORT or EXP.',
            'Products: SaaS Mobile Application, UPI QR, SoundBox, Transactional/Promotional Messages, PAN/Aadhaar/GST Verification, Late Fees Charges',
            'HSN 9971 = SaaS/Mobile App/UPI QR/Additional Users; HSN 997319 = SoundBox; HSN 998599 = SMS/App Notifications/Messages; HSN 998529 = PAN/Aadhaar/Verification; HSN 998311 = Late Fees Charges',
            'BILLING TABLE: Extract ONLY from the main billing/invoice table (sections A through G, each with Particulars, HSN, Amount, Discount, CGST, SGST, IGST columns). Do NOT extract from plan-description headers, usage-statistics sections (user count / transaction count rows), or balance-due summary rows.',
            'TAXABLE VALUE: Use the "Sub Total" shown at the end of each billing section as taxable_value. If the section shows "Amount" and "Discount on Application", taxable_value = Amount − Discount (may be zero or negative). The plan label like "Rs. 5000 per month" is a plan name — use the printed Amount column value, not a computed plan price.',
            'DISCOUNT FIELD: Set discount = the Discount column value for that billing row (e.g. if "Discount on Application = 2500", set discount=2500). If no discount row, discount = 0.',
            'TAX AMOUNTS — OVERRIDE RULES 7 AND 9: Read the EXACT CGST, SGST, IGST amounts printed in the tax columns for each billing row. Do NOT compute tax amounts from taxable × rate. Do NOT infer intra/interstate from GSTIN state codes — just read whatever amounts are printed in the CGST / SGST / IGST columns for that row.',
            'ADVANCES: "Less: Advance", "Previous Payment", or "Application Charges" in the balance-due summary section is NOT a billing row. Do not extract it as an item or use its value as any item amount or taxable value.',
            'COMBINED ROWS: Some invoices combine all services into one line (e.g. "SaaS Mobile Application, UPI QR, Transactional Messages, Late Fees Charges"). Treat as one item.',
        ],
        'few_shot_examples': examples,
        'trained_on':       TRAINING_REFS,
        'invoice_count':    len(TRAINING_REFS),
        'pass_rate_pct':    round(100*passed/max(total,1), 1),
        'trained_at':       datetime.now().isoformat(),
    }

    with open(PROFILE_FILE, 'w', encoding='utf-8') as f:
        json.dump(profile, f, indent=2, ensure_ascii=False)

    print(f"\nGolden  → {GOLDEN_FILE}")
    print(f"Profile → {PROFILE_FILE}")
    if examples:
        print(f"\nFew-shot examples ({len(examples)}):")
        for ex in examples:
            print(f"  [{ex['invoice_type']}] {ex['ref']} — {ex['output']['particulars'][:55]}")

if __name__ == '__main__':
    main()
