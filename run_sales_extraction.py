import os
import sys
import glob
import pandas as pd
from openpyxl.styles import PatternFill

# Add backend to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'backend'))
from invoice_processor import process_pdf, build_dataframes

sales_dir = r"c:\Users\yugvk\Downloads\antigravityaudit\salesinvoices"
output_path = os.path.join(sales_dir, "Sales_Output.xlsx")

pdf_files = glob.glob(os.path.join(sales_dir, "*.pdf"))

all_mains = []
all_narrations = []
all_lines = []

for pdf_path in pdf_files:
    print(f"Processing: {os.path.basename(pdf_path)}")
    res = process_pdf(pdf_path, invoice_type='Sales')
    sales_dfs, _ = build_dataframes(res)
    
    if not sales_dfs["Main"].empty:
        all_mains.append(sales_dfs["Main"])
    if not sales_dfs["Narration"].empty:
        all_narrations.append(sales_dfs["Narration"])
    if not sales_dfs["LineItems"].empty:
        all_lines.append(sales_dfs["LineItems"])

# Concatenate all dataframes
final_main = pd.concat(all_mains, ignore_index=True) if all_mains else pd.DataFrame()
final_narration = pd.concat(all_narrations, ignore_index=True) if all_narrations else pd.DataFrame()
final_lines = pd.concat(all_lines, ignore_index=True) if all_lines else pd.DataFrame()

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    final_main.to_excel(writer, sheet_name="Main", index=False)
    final_narration.to_excel(writer, sheet_name="Narration", index=False)
    final_lines.to_excel(writer, sheet_name="Line Items", index=False)
    
    # Apply styling
    workbook = writer.book
    if "Narration" in writer.sheets:
        narration_sheet = writer.sheets["Narration"]
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        for row in range(2, len(final_narration) + 2):
            narration_sheet.cell(row=row, column=1).fill = yellow_fill

print(f"Excel saved to: {output_path}")
